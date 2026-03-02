# import os
# import re
# import json
# import base64
# import argparse
# from pathlib import Path

# import pandas as pd
# import gspread
# from google.oauth2.service_account import Credentials
# from dotenv import load_dotenv
# from openpyxl import load_workbook

# # ============================================================
# # ENV LOADING (IMPORTANT FIX)
# # ------------------------------------------------------------
# # Problem you had:
# # - python-dotenv was reading the WRONG .env (repo root), not server/.env
# # Fix:
# # - explicitly load server/.env relative to this file
# # - OR allow ENV_PATH to override
# # ============================================================

# def _load_env_once():
#     # If user sets ENV_PATH in Node or system env, use it.
#     env_path_override = os.getenv("ENV_PATH", "").strip()
#     if env_path_override:
#         p = Path(env_path_override).expanduser().resolve()
#         load_dotenv(dotenv_path=p, override=True)
#         return

#     # Default: <repo_root>/server/.env, where repo_root is folder containing this file
#     # payable_TD.py is in repo root in your setup
#     here = Path(__file__).resolve().parent
#     server_env = here / "server" / ".env"
#     if server_env.exists():
#         load_dotenv(dotenv_path=server_env, override=True)
#         return

#     # Fallback: try standard load_dotenv (last resort)
#     load_dotenv(override=True)


# _load_env_once()

# # ============================================================
# # CLI ARGS
# # ============================================================

# def parse_args():
#     ap = argparse.ArgumentParser(description="Payable Trade Discount processor")
#     ap.add_argument("--input", required=True, help="Path to input file (.csv/.xlsx/.xls)")
#     ap.add_argument("--output_dir", required=True, help="Directory to write outputs")
#     ap.add_argument(
#         "--output_name",
#         default="Payable_TD.xlsx",
#         help="Output Excel file name (default: Payable_TD.xlsx)",
#     )
#     return ap.parse_args()

# # ============================================================
# # GOOGLE CLIENTS
# # ============================================================

# SCOPES = [
#     "https://www.googleapis.com/auth/spreadsheets.readonly",
#     "https://www.googleapis.com/auth/drive.readonly",
# ]

# def get_env_required(key: str) -> str:
#     v = os.getenv(key)
#     if v is None or str(v).strip() == "":
#         raise RuntimeError(f"{key} not found in env")
#     return str(v).strip()

# def get_gspread_client():
#     b64 = get_env_required("GOOGLE_SA_JSON_B64")
#     try:
#         sa_info = json.loads(base64.b64decode(b64).decode("utf-8"))
#     except Exception as e:
#         raise RuntimeError("GOOGLE_SA_JSON_B64 is not valid base64 service account json") from e
#     creds = Credentials.from_service_account_info(sa_info, scopes=SCOPES)
#     return gspread.authorize(creds)

# def get_sheet_id():
#     return get_env_required("GOOGLE_SHEET_ID")

# # ============================================================
# # HELPERS
# # ============================================================

# def parse_percent(val) -> float:
#     if val is None or val == "":
#         return 0.0
#     if isinstance(val, (int, float)):
#         v = float(val)
#         return v / 100.0 if v > 1 else v
#     s = str(val).strip()
#     if s.endswith("%"):
#         try:
#             return float(s[:-1].strip()) / 100.0
#         except:
#             return 0.0
#     try:
#         num = float(s)
#         return num / 100.0 if num > 1 else num
#     except:
#         return 0.0

# def to_num(col: pd.Series) -> pd.Series:
#     return pd.to_numeric(
#         col.astype(str).str.replace(",", "", regex=False).str.strip(),
#         errors="coerce"
#     ).fillna(0.0)

# def pick_party_col(df: pd.DataFrame) -> str:
#     if "Party Name" in df.columns:
#         return "Party Name"
#     if "Party Name1" in df.columns:
#         return "Party Name1"
#     raise KeyError("Neither 'Party Name' nor 'Party Name1' found.")

# def ensure_hsn_column(df: pd.DataFrame) -> pd.DataFrame:
#     if "HSN" in df.columns:
#         return df
#     for alt in ["HSN SAP Code", "HSN code", "HSN Code"]:
#         if alt in df.columns:
#             return df.rename(columns={alt: "HSN"})
#     return df

# def clean_hsn_series(s: pd.Series) -> pd.Series:
#     return s.astype(str).str.replace(".0", "", regex=False).str.strip()

# def clean_text(v) -> str:
#     if v is None:
#         return ""
#     s = str(v)
#     s = s.replace("\u00A0", " ")
#     s = re.sub(r"[\u200B-\u200D\uFEFF]", "", s)
#     s = s.replace("“", "").replace("”", "").replace('"', "")
#     s = re.sub(r"\s+", " ", s).strip()
#     return s

# def safe_read_input(input_file: str) -> pd.DataFrame:
#     if not os.path.exists(input_file):
#         raise FileNotFoundError(f"Input file not found: {input_file}")

#     if input_file.lower().endswith(".csv"):
#         return pd.read_csv(input_file, dtype=str)
#     if input_file.lower().endswith((".xlsx", ".xls")):
#         return pd.read_excel(input_file, dtype=str)
#     raise ValueError("Input file must be .csv or .xlsx/.xls")

# # ============================================================
# # MASTER FETCH (Margin & Billing + HSN)
# # ============================================================

# def load_masters():
#     MARGIN_SHEET_NAME = "Margin & Billing"
#     HSN_SHEET_NAME = "HSN Master"

#     gc = get_gspread_client()
#     ss = gc.open_by_key(get_sheet_id())

#     # ---- Margin & Billing
#     margin_ws = ss.worksheet(MARGIN_SHEET_NAME)
#     margin_values = margin_ws.get_all_values()
#     if len(margin_values) < 2:
#         raise ValueError("Margin & Billing sheet is empty.")

#     m_headers = margin_values[0]

#     def col(headers, names):
#         idx = {str(h).strip().lower(): i for i, h in enumerate(headers)}
#         for n in names:
#             k = str(n).strip().lower()
#             if k in idx:
#                 return idx[k]
#         return -1

#     m_party  = col(m_headers, ["party name"])
#     m_brand  = col(m_headers, ["brand"])
#     m_bill_m = col(m_headers, ["bill margin", "bill margin %"])
#     m_pmt_m  = col(m_headers, ["pmt.margin", "pmt margin", "payment margin"])
#     m_store  = col(m_headers, ["store", "store code", "store_code"])

#     if m_party == -1 or m_brand == -1 or m_bill_m == -1 or m_pmt_m == -1:
#         raise KeyError("Margin & Billing missing required columns.")

#     margin_rows = []
#     for r in margin_values[1:]:
#         party = r[m_party] if m_party < len(r) else ""
#         brand = r[m_brand] if m_brand < len(r) else ""
#         store = r[m_store] if (m_store != -1 and m_store < len(r)) else ""

#         bill_marg = r[m_bill_m] if m_bill_m < len(r) else ""
#         pmt_marg  = r[m_pmt_m] if m_pmt_m < len(r) else ""

#         if party and brand:
#             margin_rows.append({
#                 "partyNorm": str(party).lower().strip(),
#                 "brandNorm": str(brand).lower().strip(),
#                 "storeNorm": str(store).lower().strip() if store else "",
#                 "billMargin": parse_percent(bill_marg),
#                 "pmtMargin": parse_percent(pmt_marg),
#             })

#     # ---- HSN Master
#     hsn_ws = ss.worksheet(HSN_SHEET_NAME)
#     hsn_values = hsn_ws.get_all_values()
#     if len(hsn_values) < 2:
#         raise ValueError("HSN Master sheet is empty.")

#     h_headers = hsn_values[0]
#     h_code  = col(h_headers, ["hsn sap code", "hsn", "hsn code"])
#     h_below = col(h_headers, ["below 2625"])
#     h_above = col(h_headers, ["above 2625"])
#     if h_code == -1 or h_below == -1 or h_above == -1:
#         raise KeyError("HSN Master missing required columns.")

#     hsn_map = {}
#     for r in hsn_values[1:]:
#         code = r[h_code] if h_code < len(r) else ""
#         if not code:
#             continue
#         key = str(code).replace(".0", "").strip()
#         hsn_map[key] = {
#             "below": parse_percent(r[h_below]) if h_below < len(r) else 0.0,
#             "above": parse_percent(r[h_above]) if h_above < len(r) else 0.0,
#         }

#     return margin_rows, hsn_map

# # ============================================================
# # STEP-1: Store_Code
# # ============================================================

# def add_store_code(df: pd.DataFrame) -> pd.DataFrame:
#     if "Branch Name" not in df.columns:
#         raise KeyError("Column 'Branch Name' not found for Store_Code step.")
#     df = df.copy()
#     df["Store_Code"] = df["Branch Name"].astype(str).str.strip().str[:4]
#     return df

# # ============================================================
# # STEP-2: Gross / Discount / Net
# # ============================================================

# def calculate_payable_data(df: pd.DataFrame) -> pd.DataFrame:
#     df = df.copy()
#     party_col = pick_party_col(df)

#     req = [party_col, "Brand", "MRP", "RSP", "Bill Qty", "Bill Promo Amount", "Bill Promo Name"]
#     miss = [c for c in req if c not in df.columns]
#     if miss:
#         raise KeyError(f"Missing required columns for Payable step: {miss}")

#     df["MRP"] = to_num(df["MRP"])
#     df["RSP"] = to_num(df["RSP"])
#     df["Bill Qty"] = to_num(df["Bill Qty"])
#     df["Bill Promo Amount"] = to_num(df["Bill Promo Amount"])

#     party_s = df[party_col].astype(str).str.lower()
#     brand_s = df["Brand"].astype(str).str.lower()
#     promo_name_s = df["Bill Promo Name"].astype(str).str.strip().str.upper()

#     # Gross Sales
#     gross = df["MRP"] * df["Bill Qty"]
#     exception = party_s.str.contains("agilitas brands private limited", na=False) & brand_s.str.contains("lotto", na=False)
#     gross.loc[exception] = (df["RSP"] * df["Bill Qty"]).loc[exception]

#     # Discount
#     discount = df["Bill Promo Amount"].copy()
#     inc5 = promo_name_s.str.startswith("INC5", na=False)
#     discount.loc[inc5] = 0.0

#     df["Gross Sales (Sales)"] = gross
#     df["Discount (Sales)"] = discount
#     df["Net Sales (Sales)"] = gross - discount
#     return df

# # ============================================================
# # STEP-3: Billing Working
# # ============================================================

# def calculate_billing_working(df_in: pd.DataFrame, margin_rows, hsn_map) -> pd.DataFrame:
#     HSN_THRESHOLD = 2625
#     REIMB_THRESHOLD = 2500

#     df = df_in.copy()
#     party_col = pick_party_col(df)
#     df = ensure_hsn_column(df)

#     required = [party_col, "Brand", "Gross Sales (Sales)", "Bill Qty", "HSN", "Store_Code"]
#     miss = [c for c in required if c not in df.columns]
#     if miss:
#         raise KeyError(f"Missing required columns for Billing Working: {miss}")

#     qty_s = to_num(df["Bill Qty"])
#     sales_s = to_num(df["Gross Sales (Sales)"])

#     out_cols = [
#         "Billing Margin (Billing Working)",
#         "MRP Rate (Billing Working)",
#         "GST Payable % (Billing Working)",
#         "Basic Rate (Billing Working)",
#         "% of GST Reimb. (Billing Working)",
#         "Basic Value (Billing Working)",
#         "GST Reimb. (Billing Working)",
#         "Bill Value (Billing Working)",
#     ]
#     for c in out_cols:
#         if c not in df.columns:
#             df[c] = 0.0

#     party_s = df[party_col].astype(str).str.lower()
#     brand_s = df["Brand"].astype(str).str.lower()
#     store_s = df["Store_Code"].astype(str).str.lower().str.strip()
#     hsn_s = clean_hsn_series(df["HSN"])

#     for i in range(len(df)):
#         party = party_s.iat[i]
#         brand = brand_s.iat[i]
#         store = store_s.iat[i]
#         hsn = hsn_s.iat[i]

#         qty = float(qty_s.iat[i]) or 0.0
#         sales = float(sales_s.iat[i]) or 0.0

#         # Billing Margin (store-first)
#         billing_margin = 0.0
#         found = False

#         if store:
#             for m in margin_rows:
#                 if (
#                     brand == m["brandNorm"]
#                     and (m["partyNorm"] in party)
#                     and m["storeNorm"]
#                     and store == m["storeNorm"]
#                 ):
#                     billing_margin = float(m["billMargin"])
#                     found = True
#                     break

#         if not found:
#             for m in margin_rows:
#                 if (
#                     brand == m["brandNorm"]
#                     and (m["partyNorm"] in party)
#                     and (not m["storeNorm"])
#                 ):
#                     billing_margin = float(m["billMargin"])
#                     break

#         df.at[df.index[i], "Billing Margin (Billing Working)"] = billing_margin

#         # MRP Rate
#         mrp = (sales / qty) if qty else 0.0
#         df.at[df.index[i], "MRP Rate (Billing Working)"] = mrp

#         # Puma special GST logic
#         is_puma_brand = "puma" in brand
#         is_puma_party = "puma sports india pvt ltd" in party

#         gst_payable = 0.0
#         gst_reimb_perc = 0.0

#         if is_puma_brand and is_puma_party:
#             gst_payable = 0.0
#             basic_temp = mrp - (mrp * billing_margin)
#             gst_reimb_perc = 0.05 if basic_temp <= REIMB_THRESHOLD else 0.18
#         else:
#             hinfo = hsn_map.get(hsn)
#             if hinfo:
#                 gst_payable = hinfo["above"] if mrp > HSN_THRESHOLD else hinfo["below"]

#             b_rate = mrp
#             b_rate -= mrp * billing_margin
#             b_rate -= (mrp * gst_payable) / (1 + gst_payable) if (1 + gst_payable) != 0 else 0.0

#             if hinfo:
#                 gst_reimb_perc = hinfo["above"] if b_rate > REIMB_THRESHOLD else hinfo["below"]

#         df.at[df.index[i], "GST Payable % (Billing Working)"] = gst_payable

#         basic_rate = mrp
#         basic_rate -= mrp * billing_margin
#         basic_rate -= (mrp * gst_payable) / (1 + gst_payable) if (1 + gst_payable) != 0 else 0.0
#         df.at[df.index[i], "Basic Rate (Billing Working)"] = basic_rate

#         df.at[df.index[i], "% of GST Reimb. (Billing Working)"] = gst_reimb_perc

#         basic_value = basic_rate * qty
#         df.at[df.index[i], "Basic Value (Billing Working)"] = basic_value

#         gst_reimb_value = basic_value * gst_reimb_perc
#         df.at[df.index[i], "GST Reimb. (Billing Working)"] = gst_reimb_value

#         df.at[df.index[i], "Bill Value (Billing Working)"] = basic_value + gst_reimb_value

#     return df

# # ============================================================
# # STEP-4: Puma PRJ Party Name (after Billing)
# # ============================================================

# def update_puma_prj_party_name(df_in: pd.DataFrame) -> pd.DataFrame:
#     df = df_in.copy()
#     party_col = pick_party_col(df)

#     party_s = df[party_col].astype(str)
#     store_s = df["Store_Code"].astype(str)
#     brand_s = df["Brand"].astype(str)

#     party_norm = party_s.str.lower().str.strip()
#     store_norm = store_s.str.lower().str.strip()
#     brand_norm = brand_s.str.lower().str.strip()

#     mask = (
#         party_norm.str.startswith("puma sports india pvt ltd", na=False)
#         & (brand_norm == "puma")
#         & (store_norm == "prj")
#     )

#     already = party_s.str.startswith("PRJ-")
#     df.loc[mask & ~already, party_col] = "PRJ-" + party_s[mask & ~already]
#     return df

# # ============================================================
# # STEP-5: Payment Working
# # ============================================================

# def calculate_payment_working(df_in: pd.DataFrame, margin_rows, hsn_map) -> pd.DataFrame:
#     df = df_in.copy()
#     party_col = pick_party_col(df)
#     df = ensure_hsn_column(df)

#     required = [
#         party_col, "Brand", "Net Sales (Sales)", "Bill Qty",
#         "% of GST Reimb. (Billing Working)", "HSN", "Store_Code"
#     ]
#     miss = [c for c in required if c not in df.columns]
#     if miss:
#         raise KeyError(f"Missing required columns for Payment Working: {miss}")

#     net_sales_s = to_num(df["Net Sales (Sales)"])
#     qty_s = to_num(df["Bill Qty"])
#     gst_reimb_bill_s = df["% of GST Reimb. (Billing Working)"].map(parse_percent)

#     out_cols = [
#         "Payment Margin (Payment Working)",
#         "Net MRP (Payment Working)",
#         "% of GST Payable (Payment Working)",
#         "Basic Rate (Payment Working)",
#         "% of GST Reimb. (Payment Working)",
#         "Basic Value (Payment Working)",
#         "GST Reimb. (Payment Working)",
#         "Payable (Payment Working)",
#     ]
#     for c in out_cols:
#         if c not in df.columns:
#             df[c] = 0.0

#     party_s = df[party_col].astype(str).str.lower()
#     brand_s = df["Brand"].astype(str).str.lower().str.strip()
#     store_s = df["Store_Code"].astype(str).str.lower().str.strip()
#     hsn_s = clean_hsn_series(df["HSN"])

#     for i in range(len(df)):
#         party = party_s.iat[i]
#         brand = brand_s.iat[i]
#         store = store_s.iat[i]
#         hsn = hsn_s.iat[i]

#         net_sales = float(net_sales_s.iat[i]) or 0.0
#         qty = float(qty_s.iat[i]) or 0.0
#         gst_reimb_billing = float(gst_reimb_bill_s.iat[i]) if not pd.isna(gst_reimb_bill_s.iat[i]) else 0.0

#         # Payment Margin (store-first)
#         payment_margin = 0.0
#         found = False

#         if store:
#             for m in margin_rows:
#                 if (
#                     brand == m["brandNorm"]
#                     and (m["partyNorm"] in party)
#                     and m["storeNorm"]
#                     and store == m["storeNorm"]
#                 ):
#                     payment_margin = float(m["pmtMargin"])
#                     found = True
#                     break

#         if not found:
#             for m in margin_rows:
#                 if (
#                     brand == m["brandNorm"]
#                     and (m["partyNorm"] in party)
#                     and (not m["storeNorm"])
#                 ):
#                     payment_margin = float(m["pmtMargin"])
#                     break

#         df.at[df.index[i], "Payment Margin (Payment Working)"] = payment_margin

#         # Net MRP
#         net_mrp = (net_sales / qty) if qty else 0.0
#         df.at[df.index[i], "Net MRP (Payment Working)"] = net_mrp

#         # GST payable %
#         gst_payable = 0.0
#         hinfo = hsn_map.get(hsn)
#         if hinfo:
#             gst_payable = hinfo["above"] if net_mrp > 2625 else hinfo["below"]
#         df.at[df.index[i], "% of GST Payable (Payment Working)"] = gst_payable

#         # Basic Rate formula (PRJ uses new formula)
#         is_prj_row = party.strip().startswith("prj-") or store == "prj"
#         if is_prj_row:
#             after_gst = net_mrp - (net_mrp * gst_payable)
#             basic_rate = after_gst - (after_gst * payment_margin)
#         else:
#             basic_rate = net_mrp
#             basic_rate -= net_mrp * payment_margin
#             basic_rate -= (net_mrp * gst_payable) / (1 + gst_payable) if (1 + gst_payable) != 0 else 0.0

#         df.at[df.index[i], "Basic Rate (Payment Working)"] = basic_rate

#         # GST reimb % copied from billing working
#         df.at[df.index[i], "% of GST Reimb. (Payment Working)"] = gst_reimb_billing

#         # Basic value
#         basic_value = basic_rate * qty
#         df.at[df.index[i], "Basic Value (Payment Working)"] = basic_value

#         # GST reimb value
#         gst_reimb_value = basic_value * gst_reimb_billing
#         df.at[df.index[i], "GST Reimb. (Payment Working)"] = gst_reimb_value

#         # Payable
#         df.at[df.index[i], "Payable (Payment Working)"] = basic_value + gst_reimb_value

#     return df

# # ============================================================
# # STEP-6: Discount DN
# # ============================================================

# def calculate_discount_dn(df_in: pd.DataFrame) -> pd.DataFrame:
#     df = df_in.copy()

#     required = [
#         "Basic Value (Billing Working)",
#         "Basic Value (Payment Working)",
#         "GST Reimb. (Billing Working)",
#         "GST Reimb. (Payment Working)",
#         "Bill Value (Billing Working)",
#         "Payable (Payment Working)",
#     ]
#     missing = [c for c in required if c not in df.columns]
#     if missing:
#         raise KeyError(f"Missing required columns for Discount DN: {missing}")

#     basic_billing = to_num(df["Basic Value (Billing Working)"])
#     basic_payment = to_num(df["Basic Value (Payment Working)"])
#     gst_billing   = to_num(df["GST Reimb. (Billing Working)"])
#     gst_payment   = to_num(df["GST Reimb. (Payment Working)"])
#     bill_value    = to_num(df["Bill Value (Billing Working)"])
#     payable       = to_num(df["Payable (Payment Working)"])

#     df["Basic Value (Discount DN)"] = basic_billing - basic_payment
#     df["GST (Discount DN)"] = gst_billing - gst_payment
#     df["Gross-Dis (Discount DN)"] = bill_value - payable

#     return df

# # ============================================================
# # STEP-7: Final Report (MONTH+BRAND match, PUMA PRJ special, brand aliases)
# # ============================================================

# def normalize_brand(brand: str) -> str:
#     b = clean_text(brand).upper()
#     if not b:
#         return ""
#     b = re.sub(r"\bAND\b", "&", b)
#     b = re.sub(r"\s*&\s*", "&", b)
#     b = re.sub(r"\s+", " ", b).strip()

#     aliases = {
#         "GORDON & BROS": "G&B",
#         "G&B": "G&B",
#         "LAFFATTIO": "EGOSS",
#         "EGOSS LADIES": "EGOSS",
#         "EGOSS": "EGOSS",
#         "GUY SAKLER": "GUY SKALER",
#         "GUY SKALER": "GUY SKALER",
#         "LANGAUGE": "LANGUAGE",
#         "LANGUAGE": "LANGUAGE",
#         "EL CURIO": "EL CURIO",
#         "REVUP": "REVUP",
#     }
#     return aliases.get(b, b)

# def is_prj_party(party: str) -> bool:
#     return "PRJ" in clean_text(party).upper()

# def month_key_from_source(val) -> str:
#     if val is None or val == "":
#         return ""
#     dt = pd.to_datetime(val, errors="coerce")
#     if pd.isna(dt):
#         return ""
#     return f"{dt.year:04d}-{dt.month:02d}"

# def month_key_from_ded(month_text, year_text) -> str:
#     m = clean_text(month_text).upper()[:3]
#     y = re.sub(r"\D", "", clean_text(year_text))
#     if not m or not y:
#         return ""
#     MAP = {"JAN":"01","FEB":"02","MAR":"03","APR":"04","MAY":"05","JUN":"06",
#            "JUL":"07","AUG":"08","SEP":"09","OCT":"10","NOV":"11","DEC":"12"}
#     mm = MAP.get(m, "")
#     if not mm:
#         return ""
#     if len(y) == 2:
#         y = "20" + y
#     return f"{y}-{mm}"

# def month_label_from_key(key: str) -> str:
#     if not key or "-" not in key:
#         return key
#     yy, mm = key.split("-", 1)
#     try:
#         mm_i = int(mm)
#     except:
#         return key
#     full = ["January","February","March","April","May","June",
#             "July","August","September","October","November","December"]
#     if mm_i < 1 or mm_i > 12:
#         return key
#     return f"{full[mm_i-1]} - {yy}"

# def load_deduction_map():
#     gc = get_gspread_client()
#     ss = gc.open_by_key(get_sheet_id())
#     ws = ss.worksheet("Deduction")

#     values = ws.get_all_values()
#     if len(values) < 2:
#         return {}

#     headers = [str(x).strip() for x in values[0]]
#     idx = {h.lower(): i for i, h in enumerate(headers)}

#     def find(names):
#         for n in names:
#             k = str(n).strip().lower()
#             if k in idx:
#                 return idx[k]
#         return -1

#     dMonth = find(["month"])
#     dYear  = find(["year"])
#     dPart  = find(["particulars"])
#     dBrand = find(["brand"])
#     dGst   = find(["gst hold value", "gst hold"])
#     dTds   = find(["tds on purchase(jul'25 to oct'25)", "tds on purchase"])
#     dFixed = find(["fixed incentive"])
#     dStore = find(["store incentive"])
#     dRem   = find(["remarks"])

#     must = [dMonth, dYear, dPart, dBrand, dGst, dTds, dFixed, dStore]
#     if any(i == -1 for i in must):
#         raise KeyError("Deduction sheet missing required headers")

#     by_month_brand = {}
#     for r in values[1:]:
#         mkey = month_key_from_ded(
#             r[dMonth] if dMonth < len(r) else "",
#             r[dYear] if dYear < len(r) else ""
#         )
#         if not mkey:
#             continue

#         brand_norm = normalize_brand(r[dBrand] if dBrand < len(r) else "")
#         if not brand_norm:
#             continue

#         def n(col):
#             v = r[col] if col != -1 and col < len(r) else ""
#             try:
#                 return float(str(v).replace(",", "").strip() or 0)
#             except:
#                 return 0.0

#         entry = {
#             "particulars": clean_text(r[dPart] if dPart < len(r) else "").upper(),
#             "gst": n(dGst),
#             "tds": n(dTds),
#             "fixed": n(dFixed),
#             "store": n(dStore),
#             "rem": clean_text(r[dRem]) if (dRem != -1 and dRem < len(r)) else ""
#         }

#         key = f"{mkey}|{brand_norm}"
#         by_month_brand.setdefault(key, []).append(entry)

#     return by_month_brand

# def pick_deduction(ded_map, month_key: str, party_text: str, brand_text: str):
#     b_norm = normalize_brand(brand_text)
#     key = f"{month_key}|{b_norm}"
#     lst = ded_map.get(key, [])
#     if not lst:
#         return None
#     if len(lst) == 1:
#         return lst[0]

#     if b_norm == "PUMA":
#         want_prj = is_prj_party(party_text)
#         prj_row = next((e for e in lst if "PRJ" in (e.get("particulars") or "")), None)
#         non_prj = next((e for e in lst if "PRJ" not in (e.get("particulars") or "")), None)
#         return prj_row if want_prj else (non_prj or lst[0])

#     return lst[0]

# def generate_final_report(df_payable: pd.DataFrame) -> pd.DataFrame:
#     df = df_payable.copy()
#     party_col = pick_party_col(df)

#     needed = [
#         "Month", party_col, "Brand",
#         "Bill Qty", "Gross Sales (Sales)", "Discount (Sales)", "Net Sales (Sales)",
#         "Payable (Payment Working)",
#         "Basic Value (Discount DN)", "GST (Discount DN)", "Gross-Dis (Discount DN)"
#     ]
#     miss = [c for c in needed if c not in df.columns]
#     if miss:
#         raise KeyError(f"Missing required columns for FINAL report: {miss}")

#     for c in ["Bill Qty","Gross Sales (Sales)","Discount (Sales)","Net Sales (Sales)",
#               "Payable (Payment Working)","Basic Value (Discount DN)","GST (Discount DN)","Gross-Dis (Discount DN)"]:
#         df[c] = to_num(df[c])

#     df["__monthKey"] = df["Month"].map(month_key_from_source)
#     df["__monthLabel"] = df["__monthKey"].map(month_label_from_key)
#     df["__party"] = df[party_col].map(clean_text)
#     df["__brandNorm"] = df["Brand"].map(normalize_brand)

#     df = df[(df["__monthKey"] != "") & (df["__party"] != "") & (df["__brandNorm"] != "")].copy()

#     grp_cols = ["__monthKey", "__monthLabel", "__party", "__brandNorm", "Brand"]

#     agg = (
#         df.groupby(grp_cols, dropna=False)
#         .agg(
#             **{
#                 "SUM Bill Qty": ("Bill Qty", "sum"),
#                 "SUM Gross Sales": ("Gross Sales (Sales)", "sum"),
#                 "SUM Discount": ("Discount (Sales)", "sum"),
#                 "SUM Net Sales": ("Net Sales (Sales)", "sum"),
#                 "SUM Basic DN": ("Basic Value (Discount DN)", "sum"),
#                 "SUM GST DN": ("GST (Discount DN)", "sum"),
#                 "SUM Gross-Dis DN": ("Gross-Dis (Discount DN)", "sum"),
#                 "SUM of Payable before deduction (Payment Working)": ("Payable (Payment Working)", "sum"),
#             }
#         )
#         .reset_index()
#     )

#     ded_map = load_deduction_map()

#     gst_list, tds_list, fixed_list, store_list, rem_list, adj_list = [], [], [], [], [], []

#     for _, r in agg.iterrows():
#         mkey = r["__monthKey"]
#         party = r["__party"]
#         brand_display = r["Brand"]

#         ded = pick_deduction(ded_map, mkey, party, brand_display)

#         gst = ded["gst"] if ded else 0.0
#         tds = ded["tds"] if ded else 0.0
#         fixed = ded["fixed"] if ded else 0.0
#         store = ded["store"] if ded else 0.0
#         rem = ded["rem"] if ded else ""

#         total_ded = gst + tds + fixed + store
#         payable = float(r["SUM of Payable before deduction (Payment Working)"]) or 0.0
#         adjusted = payable - total_ded

#         gst_list.append(gst)
#         tds_list.append(tds)
#         fixed_list.append(fixed)
#         store_list.append(store)
#         rem_list.append(rem)
#         adj_list.append(adjusted)

#     out = pd.DataFrame({
#         "Month": agg["__monthLabel"],
#         "Party Name": agg["__party"],
#         "Brand": agg["Brand"],
#         "SUM Bill Qty": agg["SUM Bill Qty"],
#         "SUM Gross Sales": agg["SUM Gross Sales"],
#         "SUM Discount": agg["SUM Discount"],
#         "SUM Net Sales": agg["SUM Net Sales"],
#         "SUM Basic DN": agg["SUM Basic DN"],
#         "SUM GST DN": agg["SUM GST DN"],
#         "SUM Gross-Dis DN": agg["SUM Gross-Dis DN"],
#         "SUM of Payable before deduction (Payment Working)": agg["SUM of Payable before deduction (Payment Working)"],
#         "Gst Hold Value": gst_list,
#         "Tds on Purchase(Jul'25 to Oct'25)": tds_list,
#         "Fixed Incentive": fixed_list,
#         "Store Incentive": store_list,
#         "Remarks": rem_list,
#         "SUM of Payable After deduction (Payment Working)": adj_list,
#     })
#     return out

# # ============================================================
# # STEP-8: Add Section2 (Category lookup by Division+Section)
# # ============================================================

# def add_section2_from_category(df_in: pd.DataFrame) -> pd.DataFrame:
#     df = df_in.copy()

#     required = ["Division", "Section"]
#     miss = [c for c in required if c not in df.columns]
#     if miss:
#         raise KeyError(f"Missing required columns for Section2 step: {miss}")

#     gc = get_gspread_client()
#     ss = gc.open_by_key(get_sheet_id())
#     cat_ws = ss.worksheet("Category")
#     cat_values = cat_ws.get_all_values()

#     if len(cat_values) < 2:
#         df["Section2"] = ""
#         return df

#     cat_headers = cat_values[0]
#     idx = {str(h).strip().lower(): i for i, h in enumerate(cat_headers)}

#     def find(names):
#         for n in names:
#             k = str(n).strip().lower()
#             if k in idx:
#                 return idx[k]
#         return -1

#     c_div = find(["division"])
#     c_sec = find(["section"])
#     c_sec2 = find(["section2"])
#     if c_div == -1 or c_sec == -1 or c_sec2 == -1:
#         raise KeyError("Category sheet must have Division, Section, Section2")

#     cat_map = {}
#     for r in cat_values[1:]:
#         div = str(r[c_div]).strip().upper() if c_div < len(r) else ""
#         sec = str(r[c_sec]).strip().upper() if c_sec < len(r) else ""
#         sec2 = r[c_sec2] if c_sec2 < len(r) else ""
#         if not div or not sec:
#             continue
#         cat_map[f"{div}|{sec}"] = sec2  # last wins

#     div_norm = df["Division"].astype(str).str.strip().str.upper()
#     sec_norm = df["Section"].astype(str).str.strip().str.upper()
#     key_series = div_norm + "|" + sec_norm
#     df["Section2"] = key_series.map(lambda k: cat_map.get(k, ""))

#     return df

# # ============================================================
# # STEP-9: Pivot (Section2 x Store_Code => SUM Basic Value (Discount DN))
# # ============================================================

# def create_pivot_section2_store_basicdn(df_in: pd.DataFrame) -> pd.DataFrame:
#     df = df_in.copy()

#     required = ["Section2", "Store_Code", "Basic Value (Discount DN)"]
#     miss = [c for c in required if c not in df.columns]
#     if miss:
#         raise KeyError(f"Missing required columns for Pivot step: {miss}")

#     df["Basic Value (Discount DN)"] = to_num(df["Basic Value (Discount DN)"])
#     df["Section2"] = df["Section2"].astype(str)
#     df["Store_Code"] = df["Store_Code"].astype(str)

#     pivot = pd.pivot_table(
#         df,
#         index="Section2",
#         columns="Store_Code",
#         values="Basic Value (Discount DN)",
#         aggfunc="sum",
#         fill_value=0.0
#     )

#     pivot = pivot.reset_index()
#     pivot.columns.name = None
#     return pivot

# # ============================================================
# # EXCEL FORMATTING
# # ============================================================

# def format_sheet_numbers(excel_path: str, sheet_name: str, percent_cols=None, decimal_cols=None):
#     percent_cols = percent_cols or []
#     decimal_cols = decimal_cols or []

#     wb = load_workbook(excel_path)
#     ws = wb[sheet_name]

#     header_to_col = {}
#     for c in range(1, ws.max_column + 1):
#         v = ws.cell(row=1, column=c).value
#         if v is not None:
#             header_to_col[str(v)] = c

#     for name in percent_cols:
#         if name in header_to_col:
#             c = header_to_col[name]
#             for r in range(2, ws.max_row + 1):
#                 ws.cell(row=r, column=c).number_format = "0.00%"

#     for name in decimal_cols:
#         if name in header_to_col:
#             c = header_to_col[name]
#             for r in range(2, ws.max_row + 1):
#                 ws.cell(row=r, column=c).number_format = "0.############"

#     wb.save(excel_path)

# def format_pivot_all_numeric(excel_path: str, pivot_sheet: str, non_numeric_headers=("Section2",)):
#     wb = load_workbook(excel_path)
#     ws = wb[pivot_sheet]
#     headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
#     for c, h in enumerate(headers, start=1):
#         if str(h) not in set(map(str, non_numeric_headers)):
#             for r in range(2, ws.max_row + 1):
#                 ws.cell(row=r, column=c).number_format = "0.############"
#     wb.save(excel_path)

# # ============================================================
# # MAIN
# # ============================================================

# def main():
#     args = parse_args()

#     input_file = str(Path(args.input))
#     output_dir = Path(args.output_dir)
#     output_dir.mkdir(parents=True, exist_ok=True)
#     output_excel = str(output_dir / args.output_name)

#     payable_sheet = "Payable_Base_Data"
#     final_sheet = "Final"
#     pivot_sheet = "finsl Summary"

#     # ---- Read input (CSV/XLSX)
#     df = safe_read_input(input_file)

#     # ---- Filter logic (BRANDED + INHOUSE/SOR)
#     if "Inhouse / Brand" not in df.columns or "Sor / Outright" not in df.columns:
#         raise KeyError("Missing 'Inhouse / Brand' or 'Sor / Outright' columns in input file.")

#     df["Inhouse / Brand"] = df["Inhouse / Brand"].astype(str).str.strip().str.upper()
#     df["Sor / Outright"]  = df["Sor / Outright"].astype(str).str.strip().str.upper()

#     branded_df = df[df["Inhouse / Brand"] == "BRANDED"]
#     inhouse_sor_df = df[(df["Inhouse / Brand"] == "INHOUSE") & (df["Sor / Outright"] == "SOR")]
#     final_df = pd.concat([branded_df, inhouse_sor_df], ignore_index=True)

#     # ---- Load masters once
#     margin_rows, hsn_map = load_masters()

#     # Step pipeline
#     final_df = add_store_code(final_df)
#     final_df = calculate_payable_data(final_df)
#     final_df = calculate_billing_working(final_df, margin_rows, hsn_map)
#     final_df = update_puma_prj_party_name(final_df)
#     final_df = calculate_payment_working(final_df, margin_rows, hsn_map)
#     final_df = calculate_discount_dn(final_df)

#     final_report_df = generate_final_report(final_df)
#     final_df = add_section2_from_category(final_df)
#     pivot_df = create_pivot_section2_store_basicdn(final_df)

#     # ---- Write all sheets
#     with pd.ExcelWriter(output_excel, engine="openpyxl") as writer:
#         final_df.to_excel(writer, sheet_name=payable_sheet, index=False)
#         final_report_df.to_excel(writer, sheet_name=final_sheet, index=False)
#         pivot_df.to_excel(writer, sheet_name=pivot_sheet, index=False)

#     # ---- Format sheets
#     format_sheet_numbers(
#         output_excel,
#         payable_sheet,
#         percent_cols=[
#             "Billing Margin (Billing Working)",
#             "GST Payable % (Billing Working)",
#             "% of GST Reimb. (Billing Working)",
#             "Payment Margin (Payment Working)",
#             "% of GST Payable (Payment Working)",
#             "% of GST Reimb. (Payment Working)",
#         ],
#         decimal_cols=[
#             "Gross Sales (Sales)",
#             "Discount (Sales)",
#             "Net Sales (Sales)",
#             "Bill Qty",
#             "MRP Rate (Billing Working)",
#             "Basic Rate (Billing Working)",
#             "Basic Value (Billing Working)",
#             "GST Reimb. (Billing Working)",
#             "Bill Value (Billing Working)",
#             "Net MRP (Payment Working)",
#             "Basic Rate (Payment Working)",
#             "Basic Value (Payment Working)",
#             "GST Reimb. (Payment Working)",
#             "Payable (Payment Working)",
#             "Basic Value (Discount DN)",
#             "GST (Discount DN)",
#             "Gross-Dis (Discount DN)",
#         ]
#     )

#     format_sheet_numbers(
#         output_excel,
#         final_sheet,
#         decimal_cols=[
#             "SUM Bill Qty","SUM Gross Sales","SUM Discount","SUM Net Sales",
#             "SUM Basic DN","SUM GST DN","SUM Gross-Dis DN",
#             "SUM of Payable before deduction (Payment Working)",
#             "Gst Hold Value","Tds on Purchase(Jul'25 to Oct'25)",
#             "Fixed Incentive","Store Incentive",
#             "SUM of Payable After deduction (Payment Working)"
#         ]
#     )

#     format_pivot_all_numeric(output_excel, pivot_sheet, non_numeric_headers=("Section2",))

#     # ASCII-only prints (Windows safe)
#     print("DONE: Steps 1-9 completed (Payable + Final + Section2 + Pivot).")
#     print("Output:", output_excel)
#     print("Payable rows:", len(final_df))
#     print("Final rows:", len(final_report_df))
#     print("Pivot rows:", len(pivot_df))


# if __name__ == "__main__":
#     main()



import os
import re
import json
import base64
import argparse
from pathlib import Path

import pandas as pd
import gspread
from google.oauth2.service_account import Credentials
from dotenv import load_dotenv
from openpyxl import load_workbook

# ============================================================
# ENV LOADING (IMPORTANT FIX)
# ------------------------------------------------------------
# - explicitly load server/.env relative to this file
# - OR allow ENV_PATH to override
# ============================================================

def _load_env_once():
    # If user sets ENV_PATH in Node or system env, use it.
    env_path_override = os.getenv("ENV_PATH", "").strip()
    if env_path_override:
        p = Path(env_path_override).expanduser().resolve()
        load_dotenv(dotenv_path=p, override=True)
        return

    # Default: <repo_root>/server/.env
    here = Path(__file__).resolve().parent
    server_env = here / "server" / ".env"
    if server_env.exists():
        load_dotenv(dotenv_path=server_env, override=True)
        return

    # Fallback
    load_dotenv(override=True)


_load_env_once()

# ============================================================
# CLI ARGS
# ============================================================

def parse_args():
    ap = argparse.ArgumentParser(description="Payable Trade Discount processor")
    ap.add_argument("--input", required=True, help="Path to input file (.csv/.xlsx/.xls)")
    ap.add_argument("--output_dir", required=True, help="Directory to write outputs")
    ap.add_argument(
        "--output_name",
        default="Payable_TD.xlsx",
        help="Output Excel file name (default: Payable_TD.xlsx)",
    )
    return ap.parse_args()

# ============================================================
# GOOGLE CLIENTS
# ============================================================

# Need edit scope because we will update PAY_TD output tab
SCOPES = [
    "https://www.googleapis.com/auth/spreadsheets",
    "https://www.googleapis.com/auth/drive.readonly",
]

def get_env_required(key: str) -> str:
    v = os.getenv(key)
    if v is None or str(v).strip() == "":
        raise RuntimeError(f"{key} not found in env")
    return str(v).strip()

def get_gspread_client():
    b64 = get_env_required("GOOGLE_SA_JSON_B64")
    try:
        sa_info = json.loads(base64.b64decode(b64).decode("utf-8"))
    except Exception as e:
        raise RuntimeError("GOOGLE_SA_JSON_B64 is not valid base64 service account json") from e
    creds = Credentials.from_service_account_info(sa_info, scopes=SCOPES)
    return gspread.authorize(creds)

def get_sheet_id():
    # This is your "masters" spreadsheet (Margin & Billing, HSN Master, Deduction, Category...)
    return get_env_required("GOOGLE_SHEET_ID")

# ============================================================
# PAY_TD OUTPUT (UPSERT to Google Sheet)
# ============================================================

def parse_month_year_label(label: str):
    """
    Input examples:
      "December - 2025"
      "DECEMBER - 2025"
      "December-2025"
      "Dec - 25"
    Returns: ("DEC", 2025) or ("DECEMBER", 2025) depending on preference

    We'll store Month as 3-letter: JAN/FEB/.../DEC
    """
    if label is None:
        return None, None

    s = str(label).strip()
    if not s:
        return None, None

    s = re.sub(r"\s+", " ", s)

    # Try "Month - Year"
    m = re.search(r"([A-Za-z]+)\s*-\s*(\d{2,4})", s)
    if not m:
        # Try "Month Year"
        m = re.search(r"([A-Za-z]+)\s+(\d{2,4})", s)
    if not m:
        return None, None

    mon_txt = m.group(1).strip().upper()
    yy = m.group(2).strip()

    # normalize month to 3-letter
    mon3_map = {
        "JANUARY": "JAN", "JAN": "JAN",
        "FEBRUARY": "FEB", "FEB": "FEB",
        "MARCH": "MAR", "MAR": "MAR",
        "APRIL": "APR", "APR": "APR",
        "MAY": "MAY",
        "JUNE": "JUN", "JUN": "JUN",
        "JULY": "JUL", "JUL": "JUL",
        "AUGUST": "AUG", "AUG": "AUG",
        "SEPTEMBER": "SEP", "SEPT": "SEP", "SEP": "SEP",
        "OCTOBER": "OCT", "OCT": "OCT",
        "NOVEMBER": "NOV", "NOV": "NOV",
        "DECEMBER": "DEC", "DEC": "DEC",
    }
    mon3 = mon3_map.get(mon_txt, mon_txt[:3])

    year = int(yy)
    if len(yy) == 2:
        year = 2000 + year

    return mon3, year

def _coerce_existing_month_year(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    if "Month" in df.columns:
        df["Month"] = df["Month"].astype(str).str.strip().str.upper()
    if "Year" in df.columns:
        df["Year"] = pd.to_numeric(df["Year"], errors="coerce").astype("Int64")
    return df

def upsert_final_to_pay_td(final_report_df: pd.DataFrame):
    """
    Takes Final sheet DF which has column 'Month' like "December - 2025"
    Creates separate Month (JAN/FEB/...) and Year columns,
    then upserts into PAY_TD tab (replace same Month+Year, else append).
    """
    out_sheet_id = get_env_required("PAY_TD_SHEET_ID")
    out_tab_name = get_env_required("PAY_TD_TAB_NAME").strip().strip('"').strip("'")

    df_out = final_report_df.copy()

    if "Month" not in df_out.columns:
        raise KeyError("Final sheet DF must contain 'Month' column (month-year label).")

    # ✅ Rename existing label column so we can create new Month column
    df_out = df_out.rename(columns={"Month": "Month_Label"})

    # Parse Month_Label -> Month, Year
    months = []
    years = []
    for v in df_out["Month_Label"].tolist():
        m, y = parse_month_year_label(v)
        months.append(m)
        years.append(y)

    df_out.insert(0, "Year", years)
    df_out.insert(0, "Month", months)

    # Validate parsing
    if df_out["Month"].isna().any() or df_out["Year"].isna().any():
        bad = df_out[df_out["Month"].isna() | df_out["Year"].isna()].head(5)
        raise ValueError(f"Could not parse Month/Year from some rows. Examples:\n{bad[['Month_Label']]}")

    # Upsert key (this report is one month at a time)
    month_key = str(df_out.iloc[0]["Month"]).upper()
    year_key = int(df_out.iloc[0]["Year"])

    gc = get_gspread_client()
    sh = gc.open_by_key(out_sheet_id)
    ws = sh.worksheet(out_tab_name)

    values = ws.get_all_values()

    # If empty sheet, write header + data
    if not values:
        payload = [df_out.columns.tolist()] + df_out.fillna("").astype(str).values.tolist()
        ws.update(payload)
        print(f"[PAY_TD] Sheet empty. Written {len(df_out)} rows for {month_key}-{year_key}.")
        return

    header = [h.strip() for h in values[0]]
    rows = values[1:]

    # If header invalid, rebuild completely
    if not header or "Month" not in header or "Year" not in header:
        ws.clear()
        payload = [df_out.columns.tolist()] + df_out.fillna("").astype(str).values.tolist()
        ws.update(payload)
        print(f"[PAY_TD] Header missing/invalid. Rebuilt with {len(df_out)} rows for {month_key}-{year_key}.")
        return

    df_existing = pd.DataFrame(rows, columns=header)
    df_existing = _coerce_existing_month_year(df_existing)

    # Keep only columns we write (avoid mismatch)
    df_existing = df_existing.reindex(columns=df_out.columns)

    # Remove existing rows for same Month+Year
    df_filtered = df_existing[
        ~((df_existing["Month"] == month_key) & (df_existing["Year"] == year_key))
    ].copy()

    final_df = pd.concat([df_filtered, df_out], ignore_index=True)

    # Write back
    ws.clear()
    payload = [df_out.columns.tolist()] + final_df.fillna("").astype(str).values.tolist()
    ws.update(payload)

    print(f"[PAY_TD] Upsert complete for {month_key}-{year_key}. Total rows now: {len(final_df)}")


# ============================================================
# HELPERS (original)
# ============================================================

def parse_percent(val) -> float:
    if val is None or val == "":
        return 0.0
    if isinstance(val, (int, float)):
        v = float(val)
        return v / 100.0 if v > 1 else v
    s = str(val).strip()
    if s.endswith("%"):
        try:
            return float(s[:-1].strip()) / 100.0
        except:
            return 0.0
    try:
        num = float(s)
        return num / 100.0 if num > 1 else num
    except:
        return 0.0

def to_num(col: pd.Series) -> pd.Series:
    return pd.to_numeric(
        col.astype(str).str.replace(",", "", regex=False).str.strip(),
        errors="coerce"
    ).fillna(0.0)

def pick_party_col(df: pd.DataFrame) -> str:
    if "Party Name" in df.columns:
        return "Party Name"
    if "Party Name1" in df.columns:
        return "Party Name1"
    raise KeyError("Neither 'Party Name' nor 'Party Name1' found.")

def ensure_hsn_column(df: pd.DataFrame) -> pd.DataFrame:
    if "HSN" in df.columns:
        return df
    for alt in ["HSN SAP Code", "HSN code", "HSN Code"]:
        if alt in df.columns:
            return df.rename(columns={alt: "HSN"})
    return df

def clean_hsn_series(s: pd.Series) -> pd.Series:
    return s.astype(str).str.replace(".0", "", regex=False).str.strip()

def clean_text(v) -> str:
    if v is None:
        return ""
    s = str(v)
    s = s.replace("\u00A0", " ")
    s = re.sub(r"[\u200B-\u200D\uFEFF]", "", s)
    s = s.replace("“", "").replace("”", "").replace('"', "")
    s = re.sub(r"\s+", " ", s).strip()
    return s

def safe_read_input(input_file: str) -> pd.DataFrame:
    if not os.path.exists(input_file):
        raise FileNotFoundError(f"Input file not found: {input_file}")

    if input_file.lower().endswith(".csv"):
        return pd.read_csv(input_file, dtype=str)
    if input_file.lower().endswith((".xlsx", ".xls")):
        return pd.read_excel(input_file, dtype=str)
    raise ValueError("Input file must be .csv or .xlsx/.xls")

# ============================================================
# MASTER FETCH (Margin & Billing + HSN)
# ============================================================

def load_masters():
    MARGIN_SHEET_NAME = "Margin & Billing"
    HSN_SHEET_NAME = "HSN Master"

    gc = get_gspread_client()
    ss = gc.open_by_key(get_sheet_id())

    # ---- Margin & Billing
    margin_ws = ss.worksheet(MARGIN_SHEET_NAME)
    margin_values = margin_ws.get_all_values()
    if len(margin_values) < 2:
        raise ValueError("Margin & Billing sheet is empty.")

    m_headers = margin_values[0]

    def col(headers, names):
        idx = {str(h).strip().lower(): i for i, h in enumerate(headers)}
        for n in names:
            k = str(n).strip().lower()
            if k in idx:
                return idx[k]
        return -1

    m_party  = col(m_headers, ["party name"])
    m_brand  = col(m_headers, ["brand"])
    m_bill_m = col(m_headers, ["bill margin", "bill margin %"])
    m_pmt_m  = col(m_headers, ["pmt.margin", "pmt margin", "payment margin"])
    m_store  = col(m_headers, ["store", "store code", "store_code"])

    if m_party == -1 or m_brand == -1 or m_bill_m == -1 or m_pmt_m == -1:
        raise KeyError("Margin & Billing missing required columns.")

    margin_rows = []
    for r in margin_values[1:]:
        party = r[m_party] if m_party < len(r) else ""
        brand = r[m_brand] if m_brand < len(r) else ""
        store = r[m_store] if (m_store != -1 and m_store < len(r)) else ""

        bill_marg = r[m_bill_m] if m_bill_m < len(r) else ""
        pmt_marg  = r[m_pmt_m] if m_pmt_m < len(r) else ""

        if party and brand:
            margin_rows.append({
                "partyNorm": str(party).lower().strip(),
                "brandNorm": str(brand).lower().strip(),
                "storeNorm": str(store).lower().strip() if store else "",
                "billMargin": parse_percent(bill_marg),
                "pmtMargin": parse_percent(pmt_marg),
            })

    # ---- HSN Master
    hsn_ws = ss.worksheet(HSN_SHEET_NAME)
    hsn_values = hsn_ws.get_all_values()
    if len(hsn_values) < 2:
        raise ValueError("HSN Master sheet is empty.")

    h_headers = hsn_values[0]
    h_code  = col(h_headers, ["hsn sap code", "hsn", "hsn code"])
    h_below = col(h_headers, ["below 2625"])
    h_above = col(h_headers, ["above 2625"])
    if h_code == -1 or h_below == -1 or h_above == -1:
        raise KeyError("HSN Master missing required columns.")

    hsn_map = {}
    for r in hsn_values[1:]:
        code = r[h_code] if h_code < len(r) else ""
        if not code:
            continue
        key = str(code).replace(".0", "").strip()
        hsn_map[key] = {
            "below": parse_percent(r[h_below]) if h_below < len(r) else 0.0,
            "above": parse_percent(r[h_above]) if h_above < len(r) else 0.0,
        }

    return margin_rows, hsn_map

# ============================================================
# STEP-1: Store_Code
# ============================================================

def add_store_code(df: pd.DataFrame) -> pd.DataFrame:
    if "Branch Name" not in df.columns:
        raise KeyError("Column 'Branch Name' not found for Store_Code step.")
    df = df.copy()
    df["Store_Code"] = df["Branch Name"].astype(str).str.strip().str[:4]
    return df

# ============================================================
# STEP-2: Gross / Discount / Net
# ============================================================

def calculate_payable_data(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    party_col = pick_party_col(df)

    req = [party_col, "Brand", "MRP", "RSP", "Bill Qty", "Bill Promo Amount", "Bill Promo Name"]
    miss = [c for c in req if c not in df.columns]
    if miss:
        raise KeyError(f"Missing required columns for Payable step: {miss}")

    df["MRP"] = to_num(df["MRP"])
    df["RSP"] = to_num(df["RSP"])
    df["Bill Qty"] = to_num(df["Bill Qty"])
    df["Bill Promo Amount"] = to_num(df["Bill Promo Amount"])

    party_s = df[party_col].astype(str).str.lower()
    brand_s = df["Brand"].astype(str).str.lower()
    promo_name_s = df["Bill Promo Name"].astype(str).str.strip().str.upper()

    # Gross Sales
    gross = df["MRP"] * df["Bill Qty"]
    exception = party_s.str.contains("agilitas brands private limited", na=False) & brand_s.str.contains("lotto", na=False)
    gross.loc[exception] = (df["RSP"] * df["Bill Qty"]).loc[exception]

    # Discount
    discount = df["Bill Promo Amount"].copy()
    inc5 = promo_name_s.str.startswith("INC5", na=False)
    discount.loc[inc5] = 0.0

    df["Gross Sales (Sales)"] = gross
    df["Discount (Sales)"] = discount
    df["Net Sales (Sales)"] = gross - discount
    return df

# ============================================================
# STEP-3: Billing Working
# ============================================================

def calculate_billing_working(df_in: pd.DataFrame, margin_rows, hsn_map) -> pd.DataFrame:
    HSN_THRESHOLD = 2625
    REIMB_THRESHOLD = 2500

    df = df_in.copy()
    party_col = pick_party_col(df)
    df = ensure_hsn_column(df)

    required = [party_col, "Brand", "Gross Sales (Sales)", "Bill Qty", "HSN", "Store_Code"]
    miss = [c for c in required if c not in df.columns]
    if miss:
        raise KeyError(f"Missing required columns for Billing Working: {miss}")

    qty_s = to_num(df["Bill Qty"])
    sales_s = to_num(df["Gross Sales (Sales)"])

    out_cols = [
        "Billing Margin (Billing Working)",
        "MRP Rate (Billing Working)",
        "GST Payable % (Billing Working)",
        "Basic Rate (Billing Working)",
        "% of GST Reimb. (Billing Working)",
        "Basic Value (Billing Working)",
        "GST Reimb. (Billing Working)",
        "Bill Value (Billing Working)",
    ]
    for c in out_cols:
        if c not in df.columns:
            df[c] = 0.0

    party_s = df[party_col].astype(str).str.lower()
    brand_s = df["Brand"].astype(str).str.lower()
    store_s = df["Store_Code"].astype(str).str.lower().str.strip()
    hsn_s = clean_hsn_series(df["HSN"])

    for i in range(len(df)):
        party = party_s.iat[i]
        brand = brand_s.iat[i]
        store = store_s.iat[i]
        hsn = hsn_s.iat[i]

        qty = float(qty_s.iat[i]) or 0.0
        sales = float(sales_s.iat[i]) or 0.0

        # Billing Margin (store-first)
        billing_margin = 0.0
        found = False

        if store:
            for m in margin_rows:
                if (
                    brand == m["brandNorm"]
                    and (m["partyNorm"] in party)
                    and m["storeNorm"]
                    and store == m["storeNorm"]
                ):
                    billing_margin = float(m["billMargin"])
                    found = True
                    break

        if not found:
            for m in margin_rows:
                if (
                    brand == m["brandNorm"]
                    and (m["partyNorm"] in party)
                    and (not m["storeNorm"])
                ):
                    billing_margin = float(m["billMargin"])
                    break

        df.at[df.index[i], "Billing Margin (Billing Working)"] = billing_margin

        # MRP Rate
        mrp = (sales / qty) if qty else 0.0
        df.at[df.index[i], "MRP Rate (Billing Working)"] = mrp

        # Puma special GST logic
        is_puma_brand = "puma" in brand
        is_puma_party = "puma sports india pvt ltd" in party

        gst_payable = 0.0
        gst_reimb_perc = 0.0

        if is_puma_brand and is_puma_party:
            gst_payable = 0.0
            basic_temp = mrp - (mrp * billing_margin)
            gst_reimb_perc = 0.05 if basic_temp <= REIMB_THRESHOLD else 0.18
        else:
            hinfo = hsn_map.get(hsn)
            if hinfo:
                gst_payable = hinfo["above"] if mrp > HSN_THRESHOLD else hinfo["below"]

            b_rate = mrp
            b_rate -= mrp * billing_margin
            b_rate -= (mrp * gst_payable) / (1 + gst_payable) if (1 + gst_payable) != 0 else 0.0

            if hinfo:
                gst_reimb_perc = hinfo["above"] if b_rate > REIMB_THRESHOLD else hinfo["below"]

        df.at[df.index[i], "GST Payable % (Billing Working)"] = gst_payable

        basic_rate = mrp
        basic_rate -= mrp * billing_margin
        basic_rate -= (mrp * gst_payable) / (1 + gst_payable) if (1 + gst_payable) != 0 else 0.0
        df.at[df.index[i], "Basic Rate (Billing Working)"] = basic_rate

        df.at[df.index[i], "% of GST Reimb. (Billing Working)"] = gst_reimb_perc

        basic_value = basic_rate * qty
        df.at[df.index[i], "Basic Value (Billing Working)"] = basic_value

        gst_reimb_value = basic_value * gst_reimb_perc
        df.at[df.index[i], "GST Reimb. (Billing Working)"] = gst_reimb_value

        df.at[df.index[i], "Bill Value (Billing Working)"] = basic_value + gst_reimb_value

    return df

# ============================================================
# STEP-4: Puma PRJ Party Name (after Billing)
# ============================================================

def update_puma_prj_party_name(df_in: pd.DataFrame) -> pd.DataFrame:
    df = df_in.copy()
    party_col = pick_party_col(df)

    party_s = df[party_col].astype(str)
    store_s = df["Store_Code"].astype(str)
    brand_s = df["Brand"].astype(str)

    party_norm = party_s.str.lower().str.strip()
    store_norm = store_s.str.lower().str.strip()
    brand_norm = brand_s.str.lower().str.strip()

    mask = (
        party_norm.str.startswith("puma sports india pvt ltd", na=False)
        & (brand_norm == "puma")
        & (store_norm == "prj")
    )

    already = party_s.str.startswith("PRJ-")
    df.loc[mask & ~already, party_col] = "PRJ-" + party_s[mask & ~already]
    return df

# ============================================================
# STEP-5: Payment Working
# ============================================================

def calculate_payment_working(df_in: pd.DataFrame, margin_rows, hsn_map) -> pd.DataFrame:
    df = df_in.copy()
    party_col = pick_party_col(df)
    df = ensure_hsn_column(df)

    required = [
        party_col, "Brand", "Net Sales (Sales)", "Bill Qty",
        "% of GST Reimb. (Billing Working)", "HSN", "Store_Code"
    ]
    miss = [c for c in required if c not in df.columns]
    if miss:
        raise KeyError(f"Missing required columns for Payment Working: {miss}")

    net_sales_s = to_num(df["Net Sales (Sales)"])
    qty_s = to_num(df["Bill Qty"])
    gst_reimb_bill_s = df["% of GST Reimb. (Billing Working)"].map(parse_percent)

    out_cols = [
        "Payment Margin (Payment Working)",
        "Net MRP (Payment Working)",
        "% of GST Payable (Payment Working)",
        "Basic Rate (Payment Working)",
        "% of GST Reimb. (Payment Working)",
        "Basic Value (Payment Working)",
        "GST Reimb. (Payment Working)",
        "Payable (Payment Working)",
    ]
    for c in out_cols:
        if c not in df.columns:
            df[c] = 0.0

    party_s = df[party_col].astype(str).str.lower()
    brand_s = df["Brand"].astype(str).str.lower().str.strip()
    store_s = df["Store_Code"].astype(str).str.lower().str.strip()
    hsn_s = clean_hsn_series(df["HSN"])

    for i in range(len(df)):
        party = party_s.iat[i]
        brand = brand_s.iat[i]
        store = store_s.iat[i]
        hsn = hsn_s.iat[i]

        net_sales = float(net_sales_s.iat[i]) or 0.0
        qty = float(qty_s.iat[i]) or 0.0
        gst_reimb_billing = float(gst_reimb_bill_s.iat[i]) if not pd.isna(gst_reimb_bill_s.iat[i]) else 0.0

        # Payment Margin (store-first)
        payment_margin = 0.0
        found = False

        if store:
            for m in margin_rows:
                if (
                    brand == m["brandNorm"]
                    and (m["partyNorm"] in party)
                    and m["storeNorm"]
                    and store == m["storeNorm"]
                ):
                    payment_margin = float(m["pmtMargin"])
                    found = True
                    break

        if not found:
            for m in margin_rows:
                if (
                    brand == m["brandNorm"]
                    and (m["partyNorm"] in party)
                    and (not m["storeNorm"])
                ):
                    payment_margin = float(m["pmtMargin"])
                    break

        df.at[df.index[i], "Payment Margin (Payment Working)"] = payment_margin

        # Net MRP
        net_mrp = (net_sales / qty) if qty else 0.0
        df.at[df.index[i], "Net MRP (Payment Working)"] = net_mrp

        # GST payable %
        gst_payable = 0.0
        hinfo = hsn_map.get(hsn)
        if hinfo:
            gst_payable = hinfo["above"] if net_mrp > 2625 else hinfo["below"]
        df.at[df.index[i], "% of GST Payable (Payment Working)"] = gst_payable

        # Basic Rate formula (PRJ uses new formula)
        is_prj_row = party.strip().startswith("prj-") or store == "prj"
        if is_prj_row:
            after_gst = net_mrp - (net_mrp * gst_payable)
            basic_rate = after_gst - (after_gst * payment_margin)
        else:
            basic_rate = net_mrp
            basic_rate -= net_mrp * payment_margin
            basic_rate -= (net_mrp * gst_payable) / (1 + gst_payable) if (1 + gst_payable) != 0 else 0.0

        df.at[df.index[i], "Basic Rate (Payment Working)"] = basic_rate

        # GST reimb % copied from billing working
        df.at[df.index[i], "% of GST Reimb. (Payment Working)"] = gst_reimb_billing

        # Basic value
        basic_value = basic_rate * qty
        df.at[df.index[i], "Basic Value (Payment Working)"] = basic_value

        # GST reimb value
        gst_reimb_value = basic_value * gst_reimb_billing
        df.at[df.index[i], "GST Reimb. (Payment Working)"] = gst_reimb_value

        # Payable
        df.at[df.index[i], "Payable (Payment Working)"] = basic_value + gst_reimb_value

    return df

# ============================================================
# STEP-6: Discount DN
# ============================================================

def calculate_discount_dn(df_in: pd.DataFrame) -> pd.DataFrame:
    df = df_in.copy()

    required = [
        "Basic Value (Billing Working)",
        "Basic Value (Payment Working)",
        "GST Reimb. (Billing Working)",
        "GST Reimb. (Payment Working)",
        "Bill Value (Billing Working)",
        "Payable (Payment Working)",
    ]
    missing = [c for c in required if c not in df.columns]
    if missing:
        raise KeyError(f"Missing required columns for Discount DN: {missing}")

    basic_billing = to_num(df["Basic Value (Billing Working)"])
    basic_payment = to_num(df["Basic Value (Payment Working)"])
    gst_billing   = to_num(df["GST Reimb. (Billing Working)"])
    gst_payment   = to_num(df["GST Reimb. (Payment Working)"])
    bill_value    = to_num(df["Bill Value (Billing Working)"])
    payable       = to_num(df["Payable (Payment Working)"])

    df["Basic Value (Discount DN)"] = basic_billing - basic_payment
    df["GST (Discount DN)"] = gst_billing - gst_payment
    df["Gross-Dis (Discount DN)"] = bill_value - payable

    return df

# ============================================================
# STEP-7: Final Report (MONTH+BRAND match, PUMA PRJ special, brand aliases)
# ============================================================

def normalize_brand(brand: str) -> str:
    b = clean_text(brand).upper()
    if not b:
        return ""
    b = re.sub(r"\bAND\b", "&", b)
    b = re.sub(r"\s*&\s*", "&", b)
    b = re.sub(r"\s+", " ", b).strip()

    aliases = {
        "GORDON & BROS": "G&B",
        "G&B": "G&B",
        "LAFFATTIO": "EGOSS",
        "EGOSS LADIES": "EGOSS",
        "EGOSS": "EGOSS",
        "GUY SAKLER": "GUY SKALER",
        "GUY SKALER": "GUY SKALER",
        "LANGAUGE": "LANGUAGE",
        "LANGUAGE": "LANGUAGE",
        "EL CURIO": "EL CURIO",
        "REVUP": "REVUP",
    }
    return aliases.get(b, b)

def is_prj_party(party: str) -> bool:
    return "PRJ" in clean_text(party).upper()

def month_key_from_source(val) -> str:
    if val is None or val == "":
        return ""
    dt = pd.to_datetime(val, errors="coerce")
    if pd.isna(dt):
        return ""
    return f"{dt.year:04d}-{dt.month:02d}"

def month_label_from_key(key: str) -> str:
    if not key or "-" not in key:
        return key
    yy, mm = key.split("-", 1)
    try:
        mm_i = int(mm)
    except:
        return key
    full = ["January","February","March","April","May","June",
            "July","August","September","October","November","December"]
    if mm_i < 1 or mm_i > 12:
        return key
    return f"{full[mm_i-1]} - {yy}"

def month_key_from_ded(month_text, year_text) -> str:
    m = clean_text(month_text).upper()[:3]
    y = re.sub(r"\D", "", clean_text(year_text))
    if not m or not y:
        return ""
    MAP = {"JAN":"01","FEB":"02","MAR":"03","APR":"04","MAY":"05","JUN":"06",
           "JUL":"07","AUG":"08","SEP":"09","OCT":"10","NOV":"11","DEC":"12"}
    mm = MAP.get(m, "")
    if not mm:
        return ""
    if len(y) == 2:
        y = "20" + y
    return f"{y}-{mm}"

def load_deduction_map():
    gc = get_gspread_client()
    ss = gc.open_by_key(get_sheet_id())
    ws = ss.worksheet("Deduction")

    values = ws.get_all_values()
    if len(values) < 2:
        return {}

    headers = [str(x).strip() for x in values[0]]
    idx = {h.lower(): i for i, h in enumerate(headers)}

    def find(names):
        for n in names:
            k = str(n).strip().lower()
            if k in idx:
                return idx[k]
        return -1

    dMonth = find(["month"])
    dYear  = find(["year"])
    dPart  = find(["particulars"])
    dBrand = find(["brand"])
    dGst   = find(["gst hold value", "gst hold"])
    dTds   = find(["tds on purchase(jul'25 to oct'25)", "tds on purchase"])
    dFixed = find(["fixed incentive"])
    dStore = find(["store incentive"])
    dRem   = find(["remarks"])

    must = [dMonth, dYear, dPart, dBrand, dGst, dTds, dFixed, dStore]
    if any(i == -1 for i in must):
        raise KeyError("Deduction sheet missing required headers")

    by_month_brand = {}
    for r in values[1:]:
        mkey = month_key_from_ded(
            r[dMonth] if dMonth < len(r) else "",
            r[dYear] if dYear < len(r) else ""
        )
        if not mkey:
            continue

        brand_norm = normalize_brand(r[dBrand] if dBrand < len(r) else "")
        if not brand_norm:
            continue

        def n(col):
            v = r[col] if col != -1 and col < len(r) else ""
            try:
                return float(str(v).replace(",", "").strip() or 0)
            except:
                return 0.0

        entry = {
            "particulars": clean_text(r[dPart] if dPart < len(r) else "").upper(),
            "gst": n(dGst),
            "tds": n(dTds),
            "fixed": n(dFixed),
            "store": n(dStore),
            "rem": clean_text(r[dRem]) if (dRem != -1 and dRem < len(r)) else ""
        }

        key = f"{mkey}|{brand_norm}"
        by_month_brand.setdefault(key, []).append(entry)

    return by_month_brand

def pick_deduction(ded_map, month_key: str, party_text: str, brand_text: str):
    b_norm = normalize_brand(brand_text)
    key = f"{month_key}|{b_norm}"
    lst = ded_map.get(key, [])
    if not lst:
        return None
    if len(lst) == 1:
        return lst[0]

    if b_norm == "PUMA":
        want_prj = is_prj_party(party_text)
        prj_row = next((e for e in lst if "PRJ" in (e.get("particulars") or "")), None)
        non_prj = next((e for e in lst if "PRJ" not in (e.get("particulars") or "")), None)
        return prj_row if want_prj else (non_prj or lst[0])

    return lst[0]

def generate_final_report(df_payable: pd.DataFrame) -> pd.DataFrame:
    df = df_payable.copy()
    party_col = pick_party_col(df)

    needed = [
        "Month", party_col, "Brand",
        "Bill Qty", "Gross Sales (Sales)", "Discount (Sales)", "Net Sales (Sales)",
        "Payable (Payment Working)",
        "Basic Value (Discount DN)", "GST (Discount DN)", "Gross-Dis (Discount DN)"
    ]
    miss = [c for c in needed if c not in df.columns]
    if miss:
        raise KeyError(f"Missing required columns for FINAL report: {miss}")

    for c in ["Bill Qty","Gross Sales (Sales)","Discount (Sales)","Net Sales (Sales)",
              "Payable (Payment Working)","Basic Value (Discount DN)","GST (Discount DN)","Gross-Dis (Discount DN)"]:
        df[c] = to_num(df[c])

    df["__monthKey"] = df["Month"].map(month_key_from_source)
    df["__monthLabel"] = df["__monthKey"].map(month_label_from_key)
    df["__party"] = df[party_col].map(clean_text)
    df["__brandNorm"] = df["Brand"].map(normalize_brand)

    df = df[(df["__monthKey"] != "") & (df["__party"] != "") & (df["__brandNorm"] != "")].copy()

    grp_cols = ["__monthKey", "__monthLabel", "__party", "__brandNorm", "Brand"]

    agg = (
        df.groupby(grp_cols, dropna=False)
        .agg(
            **{
                "SUM Bill Qty": ("Bill Qty", "sum"),
                "SUM Gross Sales": ("Gross Sales (Sales)", "sum"),
                "SUM Discount": ("Discount (Sales)", "sum"),
                "SUM Net Sales": ("Net Sales (Sales)", "sum"),
                "SUM Basic DN": ("Basic Value (Discount DN)", "sum"),
                "SUM GST DN": ("GST (Discount DN)", "sum"),
                "SUM Gross-Dis DN": ("Gross-Dis (Discount DN)", "sum"),
                "SUM of Payable before deduction (Payment Working)": ("Payable (Payment Working)", "sum"),
            }
        )
        .reset_index()
    )

    ded_map = load_deduction_map()

    gst_list, tds_list, fixed_list, store_list, rem_list, adj_list = [], [], [], [], [], []

    for _, r in agg.iterrows():
        mkey = r["__monthKey"]
        party = r["__party"]
        brand_display = r["Brand"]

        ded = pick_deduction(ded_map, mkey, party, brand_display)

        gst = ded["gst"] if ded else 0.0
        tds = ded["tds"] if ded else 0.0
        fixed = ded["fixed"] if ded else 0.0
        store = ded["store"] if ded else 0.0
        rem = ded["rem"] if ded else ""

        total_ded = gst + tds + fixed + store
        payable = float(r["SUM of Payable before deduction (Payment Working)"]) or 0.0
        adjusted = payable - total_ded

        gst_list.append(gst)
        tds_list.append(tds)
        fixed_list.append(fixed)
        store_list.append(store)
        rem_list.append(rem)
        adj_list.append(adjusted)

    out = pd.DataFrame({
        "Month": agg["__monthLabel"],  # this will be split during upsert
        "Party Name": agg["__party"],
        "Brand": agg["Brand"],
        "SUM Bill Qty": agg["SUM Bill Qty"],
        "SUM Gross Sales": agg["SUM Gross Sales"],
        "SUM Discount": agg["SUM Discount"],
        "SUM Net Sales": agg["SUM Net Sales"],
        "SUM Basic DN": agg["SUM Basic DN"],
        "SUM GST DN": agg["SUM GST DN"],
        "SUM Gross-Dis DN": agg["SUM Gross-Dis DN"],
        "SUM of Payable before deduction (Payment Working)": agg["SUM of Payable before deduction (Payment Working)"],
        "Gst Hold Value": gst_list,
        "Tds on Purchase(Jul'25 to Oct'25)": tds_list,
        "Fixed Incentive": fixed_list,
        "Store Incentive": store_list,
        "Remarks": rem_list,
        "SUM of Payable After deduction (Payment Working)": adj_list,
    })
    return out

# ============================================================
# STEP-8: Add Section2 (Category lookup by Division+Section)
# ============================================================

def add_section2_from_category(df_in: pd.DataFrame) -> pd.DataFrame:
    df = df_in.copy()

    required = ["Division", "Section"]
    miss = [c for c in required if c not in df.columns]
    if miss:
        raise KeyError(f"Missing required columns for Section2 step: {miss}")

    gc = get_gspread_client()
    ss = gc.open_by_key(get_sheet_id())
    cat_ws = ss.worksheet("Category")
    cat_values = cat_ws.get_all_values()

    if len(cat_values) < 2:
        df["Section2"] = ""
        return df

    cat_headers = cat_values[0]
    idx = {str(h).strip().lower(): i for i, h in enumerate(cat_headers)}

    def find(names):
        for n in names:
            k = str(n).strip().lower()
            if k in idx:
                return idx[k]
        return -1

    c_div = find(["division"])
    c_sec = find(["section"])
    c_sec2 = find(["section2"])
    if c_div == -1 or c_sec == -1 or c_sec2 == -1:
        raise KeyError("Category sheet must have Division, Section, Section2")

    cat_map = {}
    for r in cat_values[1:]:
        div = str(r[c_div]).strip().upper() if c_div < len(r) else ""
        sec = str(r[c_sec]).strip().upper() if c_sec < len(r) else ""
        sec2 = r[c_sec2] if c_sec2 < len(r) else ""
        if not div or not sec:
            continue
        cat_map[f"{div}|{sec}"] = sec2  # last wins

    div_norm = df["Division"].astype(str).str.strip().str.upper()
    sec_norm = df["Section"].astype(str).str.strip().str.upper()
    key_series = div_norm + "|" + sec_norm
    df["Section2"] = key_series.map(lambda k: cat_map.get(k, ""))

    return df

# ============================================================
# STEP-9: Pivot (Section2 x Store_Code => SUM Basic Value (Discount DN))
# ============================================================

def create_pivot_section2_store_basicdn(df_in: pd.DataFrame) -> pd.DataFrame:
    df = df_in.copy()

    required = ["Section2", "Store_Code", "Basic Value (Discount DN)"]
    miss = [c for c in required if c not in df.columns]
    if miss:
        raise KeyError(f"Missing required columns for Pivot step: {miss}")

    df["Basic Value (Discount DN)"] = to_num(df["Basic Value (Discount DN)"])
    df["Section2"] = df["Section2"].astype(str)
    df["Store_Code"] = df["Store_Code"].astype(str)

    pivot = pd.pivot_table(
        df,
        index="Section2",
        columns="Store_Code",
        values="Basic Value (Discount DN)",
        aggfunc="sum",
        fill_value=0.0
    )

    pivot = pivot.reset_index()
    pivot.columns.name = None
    return pivot

# ============================================================
# EXCEL FORMATTING
# ============================================================

def format_sheet_numbers(excel_path: str, sheet_name: str, percent_cols=None, decimal_cols=None):
    percent_cols = percent_cols or []
    decimal_cols = decimal_cols or []

    wb = load_workbook(excel_path)
    ws = wb[sheet_name]

    header_to_col = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=c).value
        if v is not None:
            header_to_col[str(v)] = c

    for name in percent_cols:
        if name in header_to_col:
            c = header_to_col[name]
            for r in range(2, ws.max_row + 1):
                ws.cell(row=r, column=c).number_format = "0.00%"

    for name in decimal_cols:
        if name in header_to_col:
            c = header_to_col[name]
            for r in range(2, ws.max_row + 1):
                ws.cell(row=r, column=c).number_format = "0.############"

    wb.save(excel_path)

def format_pivot_all_numeric(excel_path: str, pivot_sheet: str, non_numeric_headers=("Section2",)):
    wb = load_workbook(excel_path)
    ws = wb[pivot_sheet]
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    for c, h in enumerate(headers, start=1):
        if str(h) not in set(map(str, non_numeric_headers)):
            for r in range(2, ws.max_row + 1):
                ws.cell(row=r, column=c).number_format = "0.############"
    wb.save(excel_path)

# ============================================================
# MAIN
# ============================================================

def main():
    args = parse_args()

    input_file = str(Path(args.input))
    output_dir = Path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    output_excel = str(output_dir / args.output_name)

    payable_sheet = "Payable_Base_Data"
    final_sheet = "Final"
    pivot_sheet = "finsl Summary"

    # ---- Read input (CSV/XLSX)
    df = safe_read_input(input_file)

    # ---- Filter logic (BRANDED + INHOUSE/SOR)
    if "Inhouse / Brand" not in df.columns or "Sor / Outright" not in df.columns:
        raise KeyError("Missing 'Inhouse / Brand' or 'Sor / Outright' columns in input file.")

    df["Inhouse / Brand"] = df["Inhouse / Brand"].astype(str).str.strip().str.upper()
    df["Sor / Outright"]  = df["Sor / Outright"].astype(str).str.strip().str.upper()

    branded_df = df[df["Inhouse / Brand"] == "BRANDED"]
    inhouse_sor_df = df[(df["Inhouse / Brand"] == "INHOUSE") & (df["Sor / Outright"] == "SOR")]
    final_df = pd.concat([branded_df, inhouse_sor_df], ignore_index=True)

    # ---- Load masters once
    margin_rows, hsn_map = load_masters()

    # Step pipeline
    final_df = add_store_code(final_df)
    final_df = calculate_payable_data(final_df)
    final_df = calculate_billing_working(final_df, margin_rows, hsn_map)
    final_df = update_puma_prj_party_name(final_df)
    final_df = calculate_payment_working(final_df, margin_rows, hsn_map)
    final_df = calculate_discount_dn(final_df)

    final_report_df = generate_final_report(final_df)
    final_df = add_section2_from_category(final_df)
    pivot_df = create_pivot_section2_store_basicdn(final_df)

    # ✅ Upsert FINAL report to PAY_TD tab with Month/Year split + replace/append logic
    upsert_final_to_pay_td(final_report_df)

    # ---- Write all sheets
    with pd.ExcelWriter(output_excel, engine="openpyxl") as writer:
        final_df.to_excel(writer, sheet_name=payable_sheet, index=False)
        final_report_df.to_excel(writer, sheet_name=final_sheet, index=False)
        pivot_df.to_excel(writer, sheet_name=pivot_sheet, index=False)

    # ---- Format sheets
    format_sheet_numbers(
        output_excel,
        payable_sheet,
        percent_cols=[
            "Billing Margin (Billing Working)",
            "GST Payable % (Billing Working)",
            "% of GST Reimb. (Billing Working)",
            "Payment Margin (Payment Working)",
            "% of GST Payable (Payment Working)",
            "% of GST Reimb. (Payment Working)",
        ],
        decimal_cols=[
            "Gross Sales (Sales)",
            "Discount (Sales)",
            "Net Sales (Sales)",
            "Bill Qty",
            "MRP Rate (Billing Working)",
            "Basic Rate (Billing Working)",
            "Basic Value (Billing Working)",
            "GST Reimb. (Billing Working)",
            "Bill Value (Billing Working)",
            "Net MRP (Payment Working)",
            "Basic Rate (Payment Working)",
            "Basic Value (Payment Working)",
            "GST Reimb. (Payment Working)",
            "Payable (Payment Working)",
            "Basic Value (Discount DN)",
            "GST (Discount DN)",
            "Gross-Dis (Discount DN)",
        ]
    )

    format_sheet_numbers(
        output_excel,
        final_sheet,
        decimal_cols=[
            "SUM Bill Qty","SUM Gross Sales","SUM Discount","SUM Net Sales",
            "SUM Basic DN","SUM GST DN","SUM Gross-Dis DN",
            "SUM of Payable before deduction (Payment Working)",
            "Gst Hold Value","Tds on Purchase(Jul'25 to Oct'25)",
            "Fixed Incentive","Store Incentive",
            "SUM of Payable After deduction (Payment Working)"
        ]
    )

    format_pivot_all_numeric(output_excel, pivot_sheet, non_numeric_headers=("Section2",))

    # ASCII-only prints (Windows safe)
    print("DONE: Steps 1-9 completed (Payable + Final + Section2 + Pivot).")
    print("Output:", output_excel)
    print("Payable rows:", len(final_df))
    print("Final rows:", len(final_report_df))
    print("Pivot rows:", len(pivot_df))


if __name__ == "__main__":
    main()
