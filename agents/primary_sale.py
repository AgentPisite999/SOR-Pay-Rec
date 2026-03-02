
# import os
# import re
# import json
# import base64
# import argparse
# from pathlib import Path
# from typing import Dict, Tuple, List

# import pandas as pd
# from dotenv import load_dotenv

# from openpyxl import load_workbook, Workbook
# from openpyxl.utils import get_column_letter
# from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

# from google.oauth2.service_account import Credentials
# from googleapiclient.discovery import build

# import gspread


# # ---------------------------
# # SHEET NAMES
# # ---------------------------
# RAW_SHEET_PRIMARY = "Primary sale raw"
# RAW_SHEET_FALLBACK = "Primary sale"
# FINAL_SHEET_NAME = "final"

# # ---------------------------
# # ENV LOADING (IMPORTANT FIX)
# # ---------------------------
# def _load_env_once():
#     env_path_override = os.getenv("ENV_PATH", "").strip()
#     if env_path_override:
#         p = Path(env_path_override).expanduser().resolve()
#         load_dotenv(dotenv_path=p, override=True)
#         return

#     here = Path(__file__).resolve().parent
#     server_env = here / "server" / ".env"
#     if server_env.exists():
#         load_dotenv(dotenv_path=server_env, override=True)
#         return

#     load_dotenv(override=True)


# _load_env_once()


# def get_env_required(key: str) -> str:
#     v = os.getenv(key)
#     if v is None or str(v).strip() == "":
#         raise RuntimeError(f"Missing env var: {key}")
#     return str(v).strip()


# # ---------------------------
# # GOOGLE MASTER (from .env)
# # ---------------------------
# GOOGLE_SA_JSON_B64 = (os.getenv("GOOGLE_SA_JSON_B64") or "").strip()
# MASTER_SPREADSHEET_ID = (os.getenv("masterSpreadsheetId") or "").strip()
# MASTER_SHEET_NAME = (os.getenv("SHEET_NAME2") or "Sheet1").strip().strip('"').strip("'")
# MASTER_SCOPES = ["https://www.googleapis.com/auth/spreadsheets.readonly"]


# # ============================================================
# # CLI ARGS
# # ============================================================
# def parse_args():
#     ap = argparse.ArgumentParser(description="Primary Sale processor")
#     ap.add_argument("--input", required=True, help="Path to input file (.xlsx/.xls)")
#     ap.add_argument("--output_dir", required=True, help="Directory to write outputs")
#     ap.add_argument("--output_name", default="Output.xlsx", help="Output Excel file name (default: Output.xlsx)")
#     return ap.parse_args()


# # =============================================================================
# # A) PRIMARY SALE RAW CREATION
# # =============================================================================
# def find_header_row(excel_path, key_col="Invoice Date", max_scan_rows=40, sheet_name=None) -> int:
#     wb = load_workbook(excel_path, read_only=True, data_only=True)
#     ws = wb[sheet_name] if sheet_name else wb.active

#     for r in range(1, max_scan_rows + 1):
#         vals = []
#         for cell in ws[r]:
#             if cell.value is not None:
#                 vals.append(str(cell.value).strip())
#         if any(v.lower() == key_col.lower() for v in vals):
#             wb.close()
#             return r

#     wb.close()
#     raise ValueError(f'Could not find header containing "{key_col}" in first {max_scan_rows} rows.')


# def build_primary_sale_raw_df(input_file: str) -> pd.DataFrame:
#     header_row_1based = find_header_row(input_file, key_col="Invoice Date", max_scan_rows=40)
#     header_row_0based = header_row_1based - 1
#     print(f"[RAW] Header found at row: {header_row_1based}")

#     df = pd.read_excel(input_file, header=header_row_0based, engine="openpyxl")
#     df = df.dropna(how="all")
#     df = df.dropna(axis=1, how="all")
#     df.columns = [str(c).strip() for c in df.columns]

#     unnamed_cols = [c for c in df.columns if str(c).lower().startswith("unnamed")]
#     df = df.drop(columns=[c for c in unnamed_cols if df[c].isna().all()], errors="ignore")

#     if "Invoice Date" not in df.columns:
#         raise KeyError(f'Column "Invoice Date" not found. Columns: {list(df.columns)}')

#     df["Invoice Date"] = pd.to_datetime(df["Invoice Date"], errors="coerce", dayfirst=True)
#     df.insert(0, "Month", df["Invoice Date"].dt.strftime("%b"))
#     df.insert(1, "Year", df["Invoice Date"].dt.strftime("%Y"))

#     return df


# def write_df_to_sheet(wb, sheet_name: str, df: pd.DataFrame):
#     if sheet_name in wb.sheetnames:
#         wb.remove(wb[sheet_name])
#     ws = wb.create_sheet(sheet_name)

#     ws.append(list(df.columns))
#     for row in df.itertuples(index=False):
#         ws.append(list(row))


# # =============================================================================
# # B) READ MASTER (Google): Groups, Customer Name
# # =============================================================================
# def get_sheets_service():
#     if not GOOGLE_SA_JSON_B64:
#         raise ValueError("Missing GOOGLE_SA_JSON_B64 in server/.env")
#     if not MASTER_SPREADSHEET_ID:
#         raise ValueError("Missing masterSpreadsheetId in server/.env")

#     sa_json = json.loads(base64.b64decode(GOOGLE_SA_JSON_B64).decode("utf-8"))
#     creds = Credentials.from_service_account_info(sa_json, scopes=MASTER_SCOPES)
#     return build("sheets", "v4", credentials=creds)


# def read_master_groups_customers() -> List[Tuple[str, str]]:
#     service = get_sheets_service()
#     resp = service.spreadsheets().values().get(
#         spreadsheetId=MASTER_SPREADSHEET_ID,
#         range=f"{MASTER_SHEET_NAME}!A:Z"
#     ).execute()

#     values = resp.get("values", [])
#     if len(values) < 2:
#         raise ValueError("Master sheet has no data.")

#     headers = [str(h).strip() for h in values[0]]
#     if "Groups" not in headers or "Customer Name" not in headers:
#         raise ValueError('Master must have headers "Groups" and "Customer Name".')

#     g_col = headers.index("Groups")
#     c_col = headers.index("Customer Name")

#     out = []
#     for r in values[1:]:
#         g = str(r[g_col]).strip() if g_col < len(r) and r[g_col] is not None else ""
#         c = str(r[c_col]).strip() if c_col < len(r) and r[c_col] is not None else ""
#         if g and c:
#             out.append((g, c))
#     return out


# # =============================================================================
# # C) FINAL SHEET TEMPLATE + formatting
# # =============================================================================
# def build_final_template(wb, master_rows: List[Tuple[str, str]], month_text: str, year_text: str):
#     if FINAL_SHEET_NAME in wb.sheetnames:
#         wb.remove(wb[FINAL_SHEET_NAME])
#     ws = wb.create_sheet(FINAL_SHEET_NAME)

#     sub_headers = ["Qty.", "Net", "Tax", "Gross", "Cost", "MRP"]
#     total_cols = 4 + 18  # A..V

#     ws.column_dimensions["A"].width = 10
#     ws.column_dimensions["B"].width = 8
#     ws.column_dimensions["C"].width = 20
#     ws.column_dimensions["D"].width = 42
#     for col in range(5, total_cols + 1):
#         ws.column_dimensions[get_column_letter(col)].width = 10

#     ws.freeze_panes = "A3"
#     ws.row_dimensions[1].height = 28
#     ws.row_dimensions[2].height = 24

#     thin = Side(style="thin", color="000000")
#     border_all = Border(left=thin, right=thin, top=thin, bottom=thin)

#     fill_sales = PatternFill("solid", fgColor="B6D7A8")
#     fill_return = PatternFill("solid", fgColor="9FC5E8")
#     fill_net = PatternFill("solid", fgColor="FFE599")

#     bold = Font(bold=True)
#     center = Alignment(horizontal="center", vertical="center")
#     mid_left = Alignment(horizontal="left", vertical="center")

#     for c in range(1, 5):
#         ws.cell(row=1, column=c, value="")

#     ws.merge_cells("E1:J1")
#     ws["E1"].value = "SALES"
#     ws["E1"].font = bold
#     ws["E1"].alignment = center
#     for cell in ws["E1:J1"][0]:
#         cell.fill = fill_sales
#         cell.font = bold
#         cell.alignment = center

#     ws.merge_cells("K1:P1")
#     ws["K1"].value = "RETURN"
#     ws["K1"].font = bold
#     ws["K1"].alignment = center
#     for cell in ws["K1:P1"][0]:
#         cell.fill = fill_return
#         cell.font = bold
#         cell.alignment = center

#     ws.merge_cells("Q1:V1")
#     ws["Q1"].value = "Net"
#     ws["Q1"].font = bold
#     ws["Q1"].alignment = center
#     for cell in ws["Q1:V1"][0]:
#         cell.fill = fill_net
#         cell.font = bold
#         cell.alignment = center

#     ws["A2"].value = "Month"
#     ws["B2"].value = "Year"
#     ws["C2"].value = "Groups"
#     ws["D2"].value = "Customer Name"
#     for addr in ["A2", "B2", "C2", "D2"]:
#         ws[addr].font = bold
#         ws[addr].alignment = center if addr in ["A2", "B2"] else mid_left

#     for i, h in enumerate(sub_headers):
#         ws.cell(row=2, column=5 + i, value=h).font = bold
#         ws.cell(row=2, column=5 + i).alignment = center
#     for i, h in enumerate(sub_headers):
#         ws.cell(row=2, column=11 + i, value=h).font = bold
#         ws.cell(row=2, column=11 + i).alignment = center
#     for i, h in enumerate(sub_headers):
#         ws.cell(row=2, column=17 + i, value=h).font = bold
#         ws.cell(row=2, column=17 + i).alignment = center

#     start_row = 3
#     for i, (grp, cust) in enumerate(master_rows):
#         r = start_row + i
#         ws.cell(row=r, column=1, value=month_text).alignment = center
#         ws.cell(row=r, column=2, value=year_text).alignment = center
#         ws.cell(row=r, column=3, value=grp)
#         ws.cell(row=r, column=4, value=cust)
#         for c in range(5, total_cols + 1):
#             ws.cell(row=r, column=c).alignment = center

#     if master_rows:
#         r = start_row
#         i = 0
#         while i < len(master_rows):
#             g = master_rows[i][0]
#             j = i + 1
#             while j < len(master_rows) and master_rows[j][0] == g:
#                 j += 1
#             block_len = j - i
#             if block_len > 1:
#                 ws.merge_cells(start_row=r, start_column=3, end_row=r + block_len - 1, end_column=3)

#             cell = ws.cell(row=r, column=3)
#             cell.font = bold
#             cell.alignment = Alignment(vertical="center")

#             r += block_len
#             i = j

#     last_row = start_row + len(master_rows) - 1 if master_rows else 2
#     gt_row = last_row + 1

#     ws.merge_cells(start_row=gt_row, start_column=1, end_row=gt_row, end_column=4)
#     gt_cell = ws.cell(row=gt_row, column=1, value="Grand Total")
#     gt_cell.font = Font(bold=True, color="FFFFFF")
#     gt_cell.fill = PatternFill("solid", fgColor="1C4587")
#     gt_cell.alignment = Alignment(horizontal="left", vertical="center")

#     for c in range(2, 5):
#         ws.cell(row=gt_row, column=c).fill = PatternFill("solid", fgColor="1C4587")

#     for rr in range(1, gt_row + 1):
#         for cc in range(1, total_cols + 1):
#             ws.cell(row=rr, column=cc).border = border_all
#             if rr <= 2:
#                 ws.cell(row=rr, column=cc).alignment = Alignment(vertical="center")

#     return ws, gt_row, start_row, total_cols


# # =============================================================================
# # D) FILL FINAL FROM RAW (aggregate + fuzzy match)
# # =============================================================================
# STOP_WORDS = {
#     "PVT", "PRIVATE", "LTD", "LIMITED", "LLP", "CO", "COMPANY",
#     "INDIA", "INTERNATIONAL", "RETAIL", "FASHION"
# }

# def norm_name(s: str) -> str:
#     s = str(s or "").upper()
#     s = re.sub(r"\([^)]*\)", " ", s)
#     s = re.sub(r"[_\-/]", " ", s)
#     s = re.sub(r"[^\w\s]", " ", s)
#     parts = [p for p in s.split() if p and p not in STOP_WORDS]
#     s = " ".join(parts)
#     s = re.sub(r"\s+", " ", s).strip()
#     return s

# def token_set(s: str) -> set:
#     return set(norm_name(s).split()) if s else set()

# def jaccard(a: set, b: set) -> float:
#     if not a or not b:
#         return 0.0
#     inter = len(a & b)
#     union = len(a | b)
#     return (inter / union) if union else 0.0

# def is_return(inv_type: str) -> bool:
#     t = str(inv_type or "").upper()
#     return "RETURN" in t

# def is_sale(inv_type: str) -> bool:
#     t = str(inv_type or "").upper()
#     return ("SALES" in t) and ("RETURN" not in t)

# def is_scrap_debtor(debtor: str) -> bool:
#     return "SCRAP" in str(debtor or "").upper()

# def resolve_col(df: pd.DataFrame, names: List[str]) -> str:
#     cols = list(df.columns)
#     lower_map = {c: str(c).strip().lower() for c in cols}
#     for c in cols:
#         h = lower_map[c]
#         for n in names:
#             n2 = n.lower()
#             if h == n2 or (n2 in h):
#                 return c
#     raise KeyError(f"Missing required column. Tried: {names}. Available: {list(df.columns)}")

# def build_aggregates_from_raw(df_raw: pd.DataFrame):
#     debtor_col = resolve_col(df_raw, ["Debtor"])
#     partner_col = resolve_col(df_raw, ["Partner"])
#     invtype_col = resolve_col(df_raw, ["Invoice Type"])

#     qty_col = resolve_col(df_raw, ["Quantity", "SUM of Quantity"])
#     net_col = resolve_col(df_raw, ["Net Amt", "SUM of Net Amt"])
#     tax_col = resolve_col(df_raw, ["Tax Gst Amt", "Tax GST Amt", "SUM of Tax", "SUM of Tax Gst", "SUM of Tax Gst Amt"])
#     gross_col = resolve_col(df_raw, ["Gross Amount", "SUM of Gross", "SUM of Gross Amount"])
#     cost_col = resolve_col(df_raw, ["Cost", "SUM of Cost"])
#     mrp_col = resolve_col(df_raw, ["MRP", "SUM of MRP"])

#     def to_num(x):
#         if x is None or x == "":
#             return 0.0
#         try:
#             return float(x)
#         except Exception:
#             return 0.0

#     agg: Dict[str, Dict[str, Dict[str, List[float]]]] = {}
#     scrap_totals = {"sale": [0.0]*6, "ret": [0.0]*6}

#     for _, row in df_raw.iterrows():
#         debtor = row.get(debtor_col, None)
#         partner = row.get(partner_col, None)
#         invtype = row.get(invtype_col, "")

#         if pd.isna(debtor) or debtor is None or str(debtor).strip() == "":
#             continue

#         vals = [
#             to_num(row.get(qty_col, 0)),
#             to_num(row.get(net_col, 0)),
#             to_num(row.get(tax_col, 0)),
#             to_num(row.get(gross_col, 0)),
#             to_num(row.get(cost_col, 0)),
#             to_num(row.get(mrp_col, 0)),
#         ]

#         bucket = "ret" if is_return(invtype) else ("sale" if is_sale(invtype) else "sale")

#         if is_scrap_debtor(debtor):
#             target = scrap_totals[bucket]
#             for k in range(6):
#                 target[k] += vals[k]
#             continue

#         g = norm_name(debtor)
#         p = norm_name(partner)

#         if g not in agg:
#             agg[g] = {}
#         if p not in agg[g]:
#             agg[g][p] = {"sale": [0.0]*6, "ret": [0.0]*6}

#         target = agg[g][p][bucket]
#         for k in range(6):
#             target[k] += vals[k]

#     return agg, scrap_totals

# def fill_final_from_raw(wb):
#     raw_name = RAW_SHEET_PRIMARY if RAW_SHEET_PRIMARY in wb.sheetnames else RAW_SHEET_FALLBACK
#     if raw_name not in wb.sheetnames:
#         raise ValueError(f'Raw sheet not found: "{RAW_SHEET_PRIMARY}" or "{RAW_SHEET_FALLBACK}"')
#     if FINAL_SHEET_NAME not in wb.sheetnames:
#         raise ValueError(f'Final sheet not found: "{FINAL_SHEET_NAME}"')

#     raw_ws = wb[raw_name]
#     raw_values = list(raw_ws.values)
#     if len(raw_values) < 2:
#         raise ValueError("Raw sheet has no data.")

#     raw_header = [str(h).strip() if h is not None else "" for h in raw_values[0]]
#     df_raw = pd.DataFrame(raw_values[1:], columns=raw_header)

#     agg, scrap_totals = build_aggregates_from_raw(df_raw)

#     ws = wb[FINAL_SHEET_NAME]

#     max_row = ws.max_row
#     grand_row = None
#     for r in range(1, max_row + 1):
#         v = ws.cell(row=r, column=1).value
#         if v and "GRAND TOTAL" in str(v).upper():
#             grand_row = r
#             break

#     data_start = 3
#     data_end = (grand_row - 1) if grand_row else max_row
#     if data_end < data_start:
#         raise ValueError("Final sheet has no customer rows.")

#     for r in range(data_start, data_end + 1):
#         for c in range(5, 23):  # E..V
#             ws.cell(row=r, column=c).value = None

#     current_group = ""
#     threshold = 0.55

#     def write_dash(row_idx: int):
#         for c in range(5, 23):
#             ws.cell(row=row_idx, column=c).value = "-"

#     for r in range(data_start, data_end + 1):
#         g_cell = ws.cell(row=r, column=3).value
#         c_name = ws.cell(row=r, column=4).value

#         if g_cell:
#             current_group = g_cell

#         if not c_name:
#             continue

#         is_scrap_row = "SCRAP" in str(current_group or "").upper()

#         if is_scrap_row:
#             sale = scrap_totals["sale"][:]
#             ret = scrap_totals["ret"][:]
#         else:
#             group_norm = norm_name(current_group)
#             cust_norm = norm_name(c_name)

#             group_map = agg.get(group_norm)
#             if not group_map:
#                 write_dash(r)
#                 continue

#             if cust_norm in group_map:
#                 sale = group_map[cust_norm]["sale"][:]
#                 ret = group_map[cust_norm]["ret"][:]
#             else:
#                 cust_tokens = token_set(c_name)
#                 best_key = None
#                 best_score = 0.0
#                 for p_key in group_map.keys():
#                     score = jaccard(cust_tokens, token_set(p_key))
#                     if score > best_score:
#                         best_score = score
#                         best_key = p_key

#                 if best_key and best_score >= threshold:
#                     sale = group_map[best_key]["sale"][:]
#                     ret = group_map[best_key]["ret"][:]
#                 else:
#                     write_dash(r)
#                     continue

#         net = [sale[i] + ret[i] for i in range(6)]

#         for i in range(6):
#             ws.cell(row=r, column=5 + i).value = sale[i]
#         for i in range(6):
#             ws.cell(row=r, column=11 + i).value = ret[i]
#         for i in range(6):
#             ws.cell(row=r, column=17 + i).value = net[i]

#     if grand_row:
#         for c in range(5, 23):
#             col_letter = get_column_letter(c)
#             ws.cell(row=grand_row, column=c).value = f"=SUM({col_letter}{data_start}:{col_letter}{data_end})"

#     print("[FINAL] Filled successfully from raw.")


# # =============================================================================
# # E) PUSH FINAL TO GOOGLE SHEET (PRIM_SALE) WITHOUT GRAND TOTAL
# # =============================================================================
# def get_gspread_client_write():
#     sa_b64 = get_env_required("GOOGLE_SA_JSON_B64")
#     try:
#         sa_info = json.loads(base64.b64decode(sa_b64).decode("utf-8"))
#     except Exception as e:
#         raise RuntimeError("GOOGLE_SA_JSON_B64 is not valid base64 service account json") from e
#     scopes = ["https://www.googleapis.com/auth/spreadsheets"]
#     creds = Credentials.from_service_account_info(sa_info, scopes=scopes)
#     return gspread.authorize(creds)


# def extract_final_df_without_grand_total(wb) -> pd.DataFrame:
#     """
#     Reads the 'final' sheet and returns a dataframe:
#     - uses row 2 as header (Month/Year/Groups/Customer Name/...).
#     - data starts from row 3
#     - stops before the row containing "Grand Total" in column A
#     - drops completely empty rows
#     """
#     if FINAL_SHEET_NAME not in wb.sheetnames:
#         raise ValueError(f'Final sheet not found: "{FINAL_SHEET_NAME}"')

#     ws = wb[FINAL_SHEET_NAME]

#     # Header is row 2 in your template
#     header_row = 2
#     headers = []
#     for c in range(1, 23):  # A..V (22 cols)
#         v = ws.cell(row=header_row, column=c).value
#         headers.append(str(v).strip() if v is not None else f"COL_{c}")

#     data = []
#     r = 3
#     while r <= ws.max_row:
#         a_val = ws.cell(row=r, column=1).value
#         if a_val and "GRAND TOTAL" in str(a_val).upper():
#             break

#         row_vals = [ws.cell(row=r, column=c).value for c in range(1, 23)]
#         # drop totally blank row
#         if any(v is not None and str(v).strip() != "" for v in row_vals):
#             data.append(row_vals)
#         r += 1

#     df = pd.DataFrame(data, columns=headers)

#     # Normalize Month/Year just in case
#     if "Month" in df.columns:
#         df["Month"] = df["Month"].astype(str).str.strip()
#     if "Year" in df.columns:
#         df["Year"] = df["Year"].astype(str).str.strip()

#     return df


# def upsert_primary_sale_to_gsheet(df_final: pd.DataFrame):
#     """
#     Upsert into PRIM_SALE tab by Month+Year:
#       - if Month+Year already exists -> remove those rows, then append new block
#       - else append at bottom
#     """
#     out_sheet_id = get_env_required("PRIM_SALE_SHEET_ID")
#     out_tab_name = get_env_required("PRIM_SALE_TAB_NAME").strip().strip('"').strip("'")

#     if df_final.empty:
#         raise ValueError("Final dataframe is empty (no rows to upload).")

#     if "Month" not in df_final.columns or "Year" not in df_final.columns:
#         raise ValueError("Final dataframe must contain Month and Year columns.")

#     # Use first row as the upsert key (one month/year per run)
#     month_key = str(df_final.iloc[0]["Month"]).strip()
#     year_key = str(df_final.iloc[0]["Year"]).strip()

#     month_key_cmp = month_key.upper()
#     year_key_cmp = year_key

#     gc = get_gspread_client_write()
#     sh = gc.open_by_key(out_sheet_id)
#     ws = sh.worksheet(out_tab_name)

#     values = ws.get_all_values()

#     # If empty -> write header + rows
#     if not values:
#         payload = [df_final.columns.tolist()] + df_final.fillna("").astype(str).values.tolist()
#         ws.update(payload)
#         print(f"[PRIM_SALE] Sheet empty. Written {len(df_final)} rows for {month_key}-{year_key}.")
#         return

#     header = [h.strip() for h in values[0]]
#     rows = values[1:]

#     # If header invalid, rebuild
#     if not header or "Month" not in header or "Year" not in header:
#         ws.clear()
#         payload = [df_final.columns.tolist()] + df_final.fillna("").astype(str).values.tolist()
#         ws.update(payload)
#         print(f"[PRIM_SALE] Header missing/invalid. Rebuilt with {len(df_final)} rows for {month_key}-{year_key}.")
#         return

#     df_existing = pd.DataFrame(rows, columns=header)

#     # Normalize for comparison
#     df_existing["Month"] = df_existing["Month"].astype(str).str.strip().str.upper()
#     df_existing["Year"] = df_existing["Year"].astype(str).str.strip()

#     # Align existing columns to new (keep it consistent)
#     df_existing = df_existing.reindex(columns=df_final.columns)

#     # Remove existing block for same Month+Year
#     df_filtered = df_existing[
#         ~((df_existing["Month"] == month_key_cmp) & (df_existing["Year"] == year_key_cmp))
#     ].copy()

#     # Ensure new df month is same normalization as sheet (not mandatory but nice)
#     df_new = df_final.copy()
#     df_new["Month"] = df_new["Month"].astype(str).str.strip().str.upper()
#     df_new["Year"] = df_new["Year"].astype(str).str.strip()

#     final_df = pd.concat([df_filtered, df_new], ignore_index=True)

#     ws.clear()
#     payload = [df_final.columns.tolist()] + final_df.fillna("").astype(str).values.tolist()
#     ws.update(payload)

#     print(f"[PRIM_SALE] Upsert complete for {month_key}-{year_key}. Total rows now: {len(final_df)}")


# # =============================================================================
# # MAIN
# # =============================================================================
# def main():
#     args = parse_args()

#     input_file = str(Path(args.input))
#     output_dir = Path(args.output_dir)
#     output_dir.mkdir(parents=True, exist_ok=True)
#     out_file = str(output_dir / args.output_name)

#     # 1) Build Primary sale raw df from input excel
#     df_raw = build_primary_sale_raw_df(input_file)

#     # Choose month/year for final sheet (use first valid row)
#     month_text = ""
#     year_text = ""
#     if "Month" in df_raw.columns and "Year" in df_raw.columns:
#         s = df_raw[["Month", "Year"]].dropna()
#         if not s.empty:
#             month_text = str(s.iloc[0]["Month"])
#             year_text = str(s.iloc[0]["Year"])

#     # 2) Load or create output workbook
#     if os.path.exists(out_file):
#         wb = load_workbook(out_file)
#     else:
#         wb = Workbook()
#         if "Sheet" in wb.sheetnames:
#             wb.remove(wb["Sheet"])

#     # 3) Write raw sheet
#     write_df_to_sheet(wb, RAW_SHEET_PRIMARY, df_raw)

#     # 4) Read master mapping from Google
#     master_rows = read_master_groups_customers()

#     # 5) Build final template
#     build_final_template(wb, master_rows, month_text, year_text)

#     # 6) Fill final from raw
#     fill_final_from_raw(wb)

#     # 7) Save file
#     wb.save(out_file)
#     print("Done. Saved:", out_file)

#     # 8) Upload FINAL to Google Sheet tab PRIM_SALE (excluding Grand Total)
#     df_final_upload = extract_final_df_without_grand_total(wb)
#     upsert_primary_sale_to_gsheet(df_final_upload)


# if __name__ == "__main__":
#     main()



import os
import re
import json
import base64
import argparse
from pathlib import Path
from typing import Dict, Tuple, List

import pandas as pd
from dotenv import load_dotenv

from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

from google.oauth2.service_account import Credentials
import gspread


# ---------------------------
# SHEET NAMES
# ---------------------------
RAW_SHEET_PRIMARY = "Primary sale raw"
RAW_SHEET_FALLBACK = "Primary sale"
FINAL_SHEET_NAME = "final"

# ---------------------------
# ENV LOADING (IMPORTANT FIX)
# ---------------------------
def _load_env_once():
    env_path_override = os.getenv("ENV_PATH", "").strip()
    if env_path_override:
        p = Path(env_path_override).expanduser().resolve()
        load_dotenv(dotenv_path=p, override=True)
        return

    here = Path(__file__).resolve().parent
    server_env = here / "server" / ".env"
    if server_env.exists():
        load_dotenv(dotenv_path=server_env, override=True)
        return

    load_dotenv(override=True)


_load_env_once()


def get_env_required(key: str) -> str:
    v = os.getenv(key)
    if v is None or str(v).strip() == "":
        raise RuntimeError(f"Missing env var: {key}")
    return str(v).strip()


# ---------------------------
# GOOGLE MASTER (from .env)
# ---------------------------
GOOGLE_SA_JSON_B64 = (os.getenv("GOOGLE_SA_JSON_B64") or "").strip()
MASTER_SPREADSHEET_ID = (os.getenv("masterSpreadsheetId") or "").strip()
MASTER_SHEET_NAME = (os.getenv("SHEET_NAME2") or "Sheet1").strip().strip('"').strip("'")


# ============================================================
# CLI ARGS
# ============================================================
def parse_args():
    ap = argparse.ArgumentParser(description="Primary Sale processor")
    ap.add_argument("--input", required=True, help="Path to input file (.xlsx/.xls)")
    ap.add_argument("--output_dir", required=True, help="Directory to write outputs")
    ap.add_argument(
        "--output_name",
        default="Output.xlsx",
        help="Output Excel file name (default: Output.xlsx)",
    )
    return ap.parse_args()


# =============================================================================
# A) PRIMARY SALE RAW CREATION
# =============================================================================
def find_header_row(excel_path, key_col="Invoice Date", max_scan_rows=40, sheet_name=None) -> int:
    wb = load_workbook(excel_path, read_only=True, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active

    for r in range(1, max_scan_rows + 1):
        vals = []
        for cell in ws[r]:
            if cell.value is not None:
                vals.append(str(cell.value).strip())
        if any(v.lower() == key_col.lower() for v in vals):
            wb.close()
            return r

    wb.close()
    raise ValueError(f'Could not find header containing "{key_col}" in first {max_scan_rows} rows.')


def build_primary_sale_raw_df(input_file: str) -> pd.DataFrame:
    header_row_1based = find_header_row(input_file, key_col="Invoice Date", max_scan_rows=40)
    header_row_0based = header_row_1based - 1
    print(f"[RAW] Header found at row: {header_row_1based}")

    df = pd.read_excel(input_file, header=header_row_0based, engine="openpyxl")
    df = df.dropna(how="all")
    df = df.dropna(axis=1, how="all")
    df.columns = [str(c).strip() for c in df.columns]

    unnamed_cols = [c for c in df.columns if str(c).lower().startswith("unnamed")]
    df = df.drop(columns=[c for c in unnamed_cols if df[c].isna().all()], errors="ignore")

    if "Invoice Date" not in df.columns:
        raise KeyError(f'Column "Invoice Date" not found. Columns: {list(df.columns)}')

    df["Invoice Date"] = pd.to_datetime(df["Invoice Date"], errors="coerce", dayfirst=True)
    df.insert(0, "Month", df["Invoice Date"].dt.strftime("%b"))
    df.insert(1, "Year", df["Invoice Date"].dt.strftime("%Y"))

    return df


def write_df_to_sheet(wb, sheet_name: str, df: pd.DataFrame):
    if sheet_name in wb.sheetnames:
        wb.remove(wb[sheet_name])
    ws = wb.create_sheet(sheet_name)

    ws.append(list(df.columns))
    for row in df.itertuples(index=False):
        ws.append(list(row))


# =============================================================================
# B) READ MASTER (Google): Groups, Customer Name  (UPDATED: gspread instead of googleapiclient)
# =============================================================================
def get_gspread_client_readonly():
    sa_b64 = get_env_required("GOOGLE_SA_JSON_B64")
    try:
        sa_info = json.loads(base64.b64decode(sa_b64).decode("utf-8"))
    except Exception as e:
        raise RuntimeError("GOOGLE_SA_JSON_B64 is not valid base64 service account json") from e

    scopes = ["https://www.googleapis.com/auth/spreadsheets.readonly"]
    creds = Credentials.from_service_account_info(sa_info, scopes=scopes)
    return gspread.authorize(creds)


def read_master_groups_customers() -> List[Tuple[str, str]]:
    if not MASTER_SPREADSHEET_ID:
        raise ValueError("Missing masterSpreadsheetId in server/.env")

    gc = get_gspread_client_readonly()
    sh = gc.open_by_key(MASTER_SPREADSHEET_ID)
    ws = sh.worksheet(MASTER_SHEET_NAME)

    values = ws.get_all_values()
    if len(values) < 2:
        raise ValueError("Master sheet has no data.")

    headers = [h.strip() for h in values[0]]
    if "Groups" not in headers or "Customer Name" not in headers:
        raise ValueError('Master must have headers "Groups" and "Customer Name".')

    g_col = headers.index("Groups")
    c_col = headers.index("Customer Name")

    out: List[Tuple[str, str]] = []
    for r in values[1:]:
        g = r[g_col].strip() if g_col < len(r) else ""
        c = r[c_col].strip() if c_col < len(r) else ""
        if g and c:
            out.append((g, c))
    return out


# =============================================================================
# C) FINAL SHEET TEMPLATE + formatting
# =============================================================================
def build_final_template(wb, master_rows: List[Tuple[str, str]], month_text: str, year_text: str):
    if FINAL_SHEET_NAME in wb.sheetnames:
        wb.remove(wb[FINAL_SHEET_NAME])
    ws = wb.create_sheet(FINAL_SHEET_NAME)

    sub_headers = ["Qty.", "Net", "Tax", "Gross", "Cost", "MRP"]
    total_cols = 4 + 18  # A..V

    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 8
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 42
    for col in range(5, total_cols + 1):
        ws.column_dimensions[get_column_letter(col)].width = 10

    ws.freeze_panes = "A3"
    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 24

    thin = Side(style="thin", color="000000")
    border_all = Border(left=thin, right=thin, top=thin, bottom=thin)

    fill_sales = PatternFill("solid", fgColor="B6D7A8")
    fill_return = PatternFill("solid", fgColor="9FC5E8")
    fill_net = PatternFill("solid", fgColor="FFE599")

    bold = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center")
    mid_left = Alignment(horizontal="left", vertical="center")

    for c in range(1, 5):
        ws.cell(row=1, column=c, value="")

    ws.merge_cells("E1:J1")
    ws["E1"].value = "SALES"
    ws["E1"].font = bold
    ws["E1"].alignment = center
    for cell in ws["E1:J1"][0]:
        cell.fill = fill_sales
        cell.font = bold
        cell.alignment = center

    ws.merge_cells("K1:P1")
    ws["K1"].value = "RETURN"
    ws["K1"].font = bold
    ws["K1"].alignment = center
    for cell in ws["K1:P1"][0]:
        cell.fill = fill_return
        cell.font = bold
        cell.alignment = center

    ws.merge_cells("Q1:V1")
    ws["Q1"].value = "Net"
    ws["Q1"].font = bold
    ws["Q1"].alignment = center
    for cell in ws["Q1:V1"][0]:
        cell.fill = fill_net
        cell.font = bold
        cell.alignment = center

    ws["A2"].value = "Month"
    ws["B2"].value = "Year"
    ws["C2"].value = "Groups"
    ws["D2"].value = "Customer Name"
    for addr in ["A2", "B2", "C2", "D2"]:
        ws[addr].font = bold
        ws[addr].alignment = center if addr in ["A2", "B2"] else mid_left

    for i, h in enumerate(sub_headers):
        ws.cell(row=2, column=5 + i, value=h).font = bold
        ws.cell(row=2, column=5 + i).alignment = center
    for i, h in enumerate(sub_headers):
        ws.cell(row=2, column=11 + i, value=h).font = bold
        ws.cell(row=2, column=11 + i).alignment = center
    for i, h in enumerate(sub_headers):
        ws.cell(row=2, column=17 + i, value=h).font = bold
        ws.cell(row=2, column=17 + i).alignment = center

    start_row = 3
    for i, (grp, cust) in enumerate(master_rows):
        r = start_row + i
        ws.cell(row=r, column=1, value=month_text).alignment = center
        ws.cell(row=r, column=2, value=year_text).alignment = center
        ws.cell(row=r, column=3, value=grp)
        ws.cell(row=r, column=4, value=cust)
        for c in range(5, total_cols + 1):
            ws.cell(row=r, column=c).alignment = center

    if master_rows:
        r = start_row
        i = 0
        while i < len(master_rows):
            g = master_rows[i][0]
            j = i + 1
            while j < len(master_rows) and master_rows[j][0] == g:
                j += 1
            block_len = j - i
            if block_len > 1:
                ws.merge_cells(start_row=r, start_column=3, end_row=r + block_len - 1, end_column=3)

            cell = ws.cell(row=r, column=3)
            cell.font = bold
            cell.alignment = Alignment(vertical="center")

            r += block_len
            i = j

    last_row = start_row + len(master_rows) - 1 if master_rows else 2
    gt_row = last_row + 1

    ws.merge_cells(start_row=gt_row, start_column=1, end_row=gt_row, end_column=4)
    gt_cell = ws.cell(row=gt_row, column=1, value="Grand Total")
    gt_cell.font = Font(bold=True, color="FFFFFF")
    gt_cell.fill = PatternFill("solid", fgColor="1C4587")
    gt_cell.alignment = Alignment(horizontal="left", vertical="center")

    for c in range(2, 5):
        ws.cell(row=gt_row, column=c).fill = PatternFill("solid", fgColor="1C4587")

    for rr in range(1, gt_row + 1):
        for cc in range(1, total_cols + 1):
            ws.cell(row=rr, column=cc).border = border_all
            if rr <= 2:
                ws.cell(row=rr, column=cc).alignment = Alignment(vertical="center")

    return ws, gt_row, start_row, total_cols


# =============================================================================
# D) FILL FINAL FROM RAW (aggregate + fuzzy match)
# =============================================================================
STOP_WORDS = {
    "PVT", "PRIVATE", "LTD", "LIMITED", "LLP", "CO", "COMPANY",
    "INDIA", "INTERNATIONAL", "RETAIL", "FASHION"
}

def norm_name(s: str) -> str:
    s = str(s or "").upper()
    s = re.sub(r"\([^)]*\)", " ", s)
    s = re.sub(r"[_\-/]", " ", s)
    s = re.sub(r"[^\w\s]", " ", s)
    parts = [p for p in s.split() if p and p not in STOP_WORDS]
    s = " ".join(parts)
    s = re.sub(r"\s+", " ", s).strip()
    return s

def token_set(s: str) -> set:
    return set(norm_name(s).split()) if s else set()

def jaccard(a: set, b: set) -> float:
    if not a or not b:
        return 0.0
    inter = len(a & b)
    union = len(a | b)
    return (inter / union) if union else 0.0

def is_return(inv_type: str) -> bool:
    t = str(inv_type or "").upper()
    return "RETURN" in t

def is_sale(inv_type: str) -> bool:
    t = str(inv_type or "").upper()
    return ("SALES" in t) and ("RETURN" not in t)

def is_scrap_debtor(debtor: str) -> bool:
    return "SCRAP" in str(debtor or "").upper()

def resolve_col(df: pd.DataFrame, names: List[str]) -> str:
    cols = list(df.columns)
    lower_map = {c: str(c).strip().lower() for c in cols}
    for c in cols:
        h = lower_map[c]
        for n in names:
            n2 = n.lower()
            if h == n2 or (n2 in h):
                return c
    raise KeyError(f"Missing required column. Tried: {names}. Available: {list(df.columns)}")

def build_aggregates_from_raw(df_raw: pd.DataFrame):
    debtor_col = resolve_col(df_raw, ["Debtor"])
    partner_col = resolve_col(df_raw, ["Partner"])
    invtype_col = resolve_col(df_raw, ["Invoice Type"])

    qty_col = resolve_col(df_raw, ["Quantity", "SUM of Quantity"])
    net_col = resolve_col(df_raw, ["Net Amt", "SUM of Net Amt"])
    tax_col = resolve_col(df_raw, ["Tax Gst Amt", "Tax GST Amt", "SUM of Tax", "SUM of Tax Gst", "SUM of Tax Gst Amt"])
    gross_col = resolve_col(df_raw, ["Gross Amount", "SUM of Gross", "SUM of Gross Amount"])
    cost_col = resolve_col(df_raw, ["Cost", "SUM of Cost"])
    mrp_col = resolve_col(df_raw, ["MRP", "SUM of MRP"])

    def to_num(x):
        if x is None or x == "":
            return 0.0
        try:
            return float(x)
        except Exception:
            return 0.0

    agg: Dict[str, Dict[str, Dict[str, List[float]]]] = {}
    scrap_totals = {"sale": [0.0]*6, "ret": [0.0]*6}

    for _, row in df_raw.iterrows():
        debtor = row.get(debtor_col, None)
        partner = row.get(partner_col, None)
        invtype = row.get(invtype_col, "")

        if pd.isna(debtor) or debtor is None or str(debtor).strip() == "":
            continue

        vals = [
            to_num(row.get(qty_col, 0)),
            to_num(row.get(net_col, 0)),
            to_num(row.get(tax_col, 0)),
            to_num(row.get(gross_col, 0)),
            to_num(row.get(cost_col, 0)),
            to_num(row.get(mrp_col, 0)),
        ]

        bucket = "ret" if is_return(invtype) else ("sale" if is_sale(invtype) else "sale")

        if is_scrap_debtor(debtor):
            target = scrap_totals[bucket]
            for k in range(6):
                target[k] += vals[k]
            continue

        g = norm_name(debtor)
        p = norm_name(partner)

        if g not in agg:
            agg[g] = {}
        if p not in agg[g]:
            agg[g][p] = {"sale": [0.0]*6, "ret": [0.0]*6}

        target = agg[g][p][bucket]
        for k in range(6):
            target[k] += vals[k]

    return agg, scrap_totals

def fill_final_from_raw(wb):
    raw_name = RAW_SHEET_PRIMARY if RAW_SHEET_PRIMARY in wb.sheetnames else RAW_SHEET_FALLBACK
    if raw_name not in wb.sheetnames:
        raise ValueError(f'Raw sheet not found: "{RAW_SHEET_PRIMARY}" or "{RAW_SHEET_FALLBACK}"')
    if FINAL_SHEET_NAME not in wb.sheetnames:
        raise ValueError(f'Final sheet not found: "{FINAL_SHEET_NAME}"')

    raw_ws = wb[raw_name]
    raw_values = list(raw_ws.values)
    if len(raw_values) < 2:
        raise ValueError("Raw sheet has no data.")

    raw_header = [str(h).strip() if h is not None else "" for h in raw_values[0]]
    df_raw = pd.DataFrame(raw_values[1:], columns=raw_header)

    agg, scrap_totals = build_aggregates_from_raw(df_raw)

    ws = wb[FINAL_SHEET_NAME]

    max_row = ws.max_row
    grand_row = None
    for r in range(1, max_row + 1):
        v = ws.cell(row=r, column=1).value
        if v and "GRAND TOTAL" in str(v).upper():
            grand_row = r
            break

    data_start = 3
    data_end = (grand_row - 1) if grand_row else max_row
    if data_end < data_start:
        raise ValueError("Final sheet has no customer rows.")

    for r in range(data_start, data_end + 1):
        for c in range(5, 23):  # E..V
            ws.cell(row=r, column=c).value = None

    current_group = ""
    threshold = 0.55

    def write_dash(row_idx: int):
        for c in range(5, 23):
            ws.cell(row=row_idx, column=c).value = "-"

    for r in range(data_start, data_end + 1):
        g_cell = ws.cell(row=r, column=3).value
        c_name = ws.cell(row=r, column=4).value

        if g_cell:
            current_group = g_cell

        if not c_name:
            continue

        is_scrap_row = "SCRAP" in str(current_group or "").upper()

        if is_scrap_row:
            sale = scrap_totals["sale"][:]
            ret = scrap_totals["ret"][:]
        else:
            group_norm = norm_name(current_group)
            cust_norm = norm_name(c_name)

            group_map = agg.get(group_norm)
            if not group_map:
                write_dash(r)
                continue

            if cust_norm in group_map:
                sale = group_map[cust_norm]["sale"][:]
                ret = group_map[cust_norm]["ret"][:]
            else:
                cust_tokens = token_set(c_name)
                best_key = None
                best_score = 0.0
                for p_key in group_map.keys():
                    score = jaccard(cust_tokens, token_set(p_key))
                    if score > best_score:
                        best_score = score
                        best_key = p_key

                if best_key and best_score >= threshold:
                    sale = group_map[best_key]["sale"][:]
                    ret = group_map[best_key]["ret"][:]
                else:
                    write_dash(r)
                    continue

        net = [sale[i] + ret[i] for i in range(6)]

        for i in range(6):
            ws.cell(row=r, column=5 + i).value = sale[i]
        for i in range(6):
            ws.cell(row=r, column=11 + i).value = ret[i]
        for i in range(6):
            ws.cell(row=r, column=17 + i).value = net[i]

    if grand_row:
        for c in range(5, 23):
            col_letter = get_column_letter(c)
            ws.cell(row=grand_row, column=c).value = f"=SUM({col_letter}{data_start}:{col_letter}{data_end})"

    print("[FINAL] Filled successfully from raw.")


# =============================================================================
# E) PUSH FINAL TO GOOGLE SHEET (PRIM_SALE) WITHOUT GRAND TOTAL
# =============================================================================
def get_gspread_client_write():
    sa_b64 = get_env_required("GOOGLE_SA_JSON_B64")
    try:
        sa_info = json.loads(base64.b64decode(sa_b64).decode("utf-8"))
    except Exception as e:
        raise RuntimeError("GOOGLE_SA_JSON_B64 is not valid base64 service account json") from e
    scopes = ["https://www.googleapis.com/auth/spreadsheets"]
    creds = Credentials.from_service_account_info(sa_info, scopes=scopes)
    return gspread.authorize(creds)


def extract_final_df_without_grand_total(wb) -> pd.DataFrame:
    if FINAL_SHEET_NAME not in wb.sheetnames:
        raise ValueError(f'Final sheet not found: "{FINAL_SHEET_NAME}"')

    ws = wb[FINAL_SHEET_NAME]

    header_row = 2
    headers = []
    for c in range(1, 23):  # A..V
        v = ws.cell(row=header_row, column=c).value
        headers.append(str(v).strip() if v is not None else f"COL_{c}")

    data = []
    r = 3
    while r <= ws.max_row:
        a_val = ws.cell(row=r, column=1).value
        if a_val and "GRAND TOTAL" in str(a_val).upper():
            break

        row_vals = [ws.cell(row=r, column=c).value for c in range(1, 23)]
        if any(v is not None and str(v).strip() != "" for v in row_vals):
            data.append(row_vals)
        r += 1

    df = pd.DataFrame(data, columns=headers)

    if "Month" in df.columns:
        df["Month"] = df["Month"].astype(str).str.strip()
    if "Year" in df.columns:
        df["Year"] = df["Year"].astype(str).str.strip()

    return df


def upsert_primary_sale_to_gsheet(df_final: pd.DataFrame):
    out_sheet_id = get_env_required("PRIM_SALE_SHEET_ID")
    out_tab_name = get_env_required("PRIM_SALE_TAB_NAME").strip().strip('"').strip("'")

    if df_final.empty:
        raise ValueError("Final dataframe is empty (no rows to upload).")

    if "Month" not in df_final.columns or "Year" not in df_final.columns:
        raise ValueError("Final dataframe must contain Month and Year columns.")

    month_key = str(df_final.iloc[0]["Month"]).strip()
    year_key = str(df_final.iloc[0]["Year"]).strip()

    month_key_cmp = month_key.upper()
    year_key_cmp = year_key

    gc = get_gspread_client_write()
    sh = gc.open_by_key(out_sheet_id)
    ws = sh.worksheet(out_tab_name)

    values = ws.get_all_values()

    if not values:
        payload = [df_final.columns.tolist()] + df_final.fillna("").astype(str).values.tolist()
        ws.update(payload)
        print(f"[PRIM_SALE] Sheet empty. Written {len(df_final)} rows for {month_key}-{year_key}.")
        return

    header = [h.strip() for h in values[0]]
    rows = values[1:]

    if not header or "Month" not in header or "Year" not in header:
        ws.clear()
        payload = [df_final.columns.tolist()] + df_final.fillna("").astype(str).values.tolist()
        ws.update(payload)
        print(f"[PRIM_SALE] Header missing/invalid. Rebuilt with {len(df_final)} rows for {month_key}-{year_key}.")
        return

    df_existing = pd.DataFrame(rows, columns=header)

    df_existing["Month"] = df_existing["Month"].astype(str).str.strip().str.upper()
    df_existing["Year"] = df_existing["Year"].astype(str).str.strip()

    df_existing = df_existing.reindex(columns=df_final.columns)

    df_filtered = df_existing[
        ~((df_existing["Month"] == month_key_cmp) & (df_existing["Year"] == year_key_cmp))
    ].copy()

    df_new = df_final.copy()
    df_new["Month"] = df_new["Month"].astype(str).str.strip().str.upper()
    df_new["Year"] = df_new["Year"].astype(str).str.strip()

    final_df = pd.concat([df_filtered, df_new], ignore_index=True)

    ws.clear()
    payload = [df_final.columns.tolist()] + final_df.fillna("").astype(str).values.tolist()
    ws.update(payload)

    print(f"[PRIM_SALE] Upsert complete for {month_key}-{year_key}. Total rows now: {len(final_df)}")


# =============================================================================
# MAIN
# =============================================================================
def main():
    args = parse_args()

    input_file = str(Path(args.input))
    output_dir = Path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    out_file = str(output_dir / args.output_name)

    df_raw = build_primary_sale_raw_df(input_file)

    month_text = ""
    year_text = ""
    if "Month" in df_raw.columns and "Year" in df_raw.columns:
        s = df_raw[["Month", "Year"]].dropna()
        if not s.empty:
            month_text = str(s.iloc[0]["Month"])
            year_text = str(s.iloc[0]["Year"])

    if os.path.exists(out_file):
        wb = load_workbook(out_file)
    else:
        wb = Workbook()
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])

    write_df_to_sheet(wb, RAW_SHEET_PRIMARY, df_raw)

    master_rows = read_master_groups_customers()
    build_final_template(wb, master_rows, month_text, year_text)
    fill_final_from_raw(wb)

    wb.save(out_file)
    print("Done. Saved:", out_file)

    df_final_upload = extract_final_df_without_grand_total(wb)
    upsert_primary_sale_to_gsheet(df_final_upload)


if __name__ == "__main__":
    main()
